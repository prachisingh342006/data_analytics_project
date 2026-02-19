import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import os

DATA_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Student_performance_data _.csv")
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Student_Early_Warning_Dashboard.xlsx")

df = pd.read_csv(DATA_PATH)
N = len(df)
DATA_END_ROW = N + 1

grade_map = {0: "A", 1: "B", 2: "C", 3: "D", 4: "F"}
df["GradeLetter"] = df["GradeClass"].map(grade_map)
df["RiskScore"] = (
    (1 - df["GPA"] / 4.0) * 35 +
    (df["Absences"] / df["Absences"].max()) * 25 +
    (1 - df["StudyTimeWeekly"] / df["StudyTimeWeekly"].max()) * 20 +
    (1 - df["ParentalSupport"] / 4.0) * 10 +
    (df["GradeClass"] / 4.0) * 10
).round(2)
df["RiskCategory"] = pd.cut(df["RiskScore"], bins=[0, 30, 55, 75, 100],
    labels=["Low", "Medium", "High", "Critical"], include_lowest=True)

# STYLES
DARK_BLUE = "1B2A4A"
MED_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
ACCENT_GOLD = "E8A838"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
RED_FILL = "FF4444"
ORANGE_FILL = "FF9944"
YELLOW_FILL = "FFD966"
GREEN_FILL = "70AD47"

hdr_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
hdr_fill = PatternFill("solid", fgColor=DARK_BLUE)
sub_hdr_font = Font(name="Calibri", bold=True, size=10, color=WHITE)
sub_hdr_fill = PatternFill("solid", fgColor=MED_BLUE)
title_font = Font(name="Calibri", bold=True, size=16, color=DARK_BLUE)
subtitle_font = Font(name="Calibri", bold=True, size=12, color=MED_BLUE)
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"))

pct_fmt = '0.0%'
num_fmt = '#,##0'
dec_fmt = '0.00'

def sc(ws, row, col, val, font=normal_font, fill=None, align=None, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font
    c.border = thin_border
    if fill: c.fill = fill
    if align: c.alignment = align
    if fmt: c.number_format = fmt
    return c

def write_hdr(ws, row, col_start, values):
    for i, v in enumerate(values):
        sc(ws, row, col_start+i, v, font=hdr_font, fill=hdr_fill,
           align=Alignment(horizontal="center", vertical="center"))

def sheet_title(ws, title, row=1, merge_to=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_to)
    c = ws.cell(row=row, column=1, value=title)
    c.font = title_font
    c.alignment = Alignment(horizontal="left", vertical="center")

def section_label(ws, text, row, col=1, merge_to=6):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    c = ws.cell(row=row, column=col, value=text)
    c.font = subtitle_font

def stripe(ws, sr, er, sc_col, ec):
    for r in range(sr, er+1):
        if (r - sr) % 2 == 1:
            for cc in range(sc_col, ec+1):
                ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=LIGHT_GRAY)

# ============================================================
wb = Workbook()

# HIDDEN DATA SHEET
ws_data = wb.active
ws_data.title = "Data"
headers_data = ["StudentID","Age","Gender","Ethnicity","ParentalEducation",
    "StudyTimeWeekly","Absences","Tutoring","ParentalSupport",
    "Extracurricular","Sports","Music","Volunteering",
    "GPA","GradeClass","GradeLetter","RiskScore","RiskCategory"]
for ci, h in enumerate(headers_data, 1):
    ws_data.cell(row=1, column=ci, value=h)
for ri, row_data in enumerate(df.itertuples(index=False), 2):
    vals = list(row_data)
    for ci, val in enumerate(vals, 1):
        if isinstance(val, (np.integer,)): val = int(val)
        elif isinstance(val, (np.floating,)): val = float(val)
        ws_data.cell(row=ri, column=ci, value=val)
ws_data.sheet_state = "hidden"

COL = {h: get_column_letter(i+1) for i, h in enumerate(headers_data)}
DER = DATA_END_ROW

# ============================================================
# PAGE 1: ACADEMIC OVERVIEW
# ============================================================
ws1 = wb.create_sheet("1. Academic Overview")
sheet_title(ws1, "Academic Overview Dashboard", row=1, merge_to=12)

# KPI labels row 3
kpi_labels = ["Total Students", "Average GPA", "Pass Rate", "Fail Rate", "Avg Study Hrs", "Avg Absences"]
for i, lbl in enumerate(kpi_labels):
    col_s = 2 + i*2
    ws1.merge_cells(start_row=3, start_column=col_s, end_row=3, end_column=col_s+1)
    sc(ws1, 3, col_s, lbl, font=sub_hdr_font, fill=sub_hdr_fill, align=Alignment(horizontal="center"))

# KPI formulas row 4
kpi_formulas = [
    ('=COUNTA(Data!A2:A' + str(DER) + ')', num_fmt),
    ('=ROUND(AVERAGE(Data!N2:N' + str(DER) + '),2)', dec_fmt),
    ('=ROUND(COUNTIF(Data!O2:O' + str(DER) + ',"<4")/COUNTA(Data!O2:O' + str(DER) + ')*100,1)', '0.0'),
    ('=ROUND(COUNTIF(Data!O2:O' + str(DER) + ',"=4")/COUNTA(Data!O2:O' + str(DER) + ')*100,1)', '0.0'),
    ('=ROUND(AVERAGE(Data!F2:F' + str(DER) + '),1)', '0.0'),
    ('=ROUND(AVERAGE(Data!G2:G' + str(DER) + '),1)', '0.0'),
]
big_font = Font(name="Calibri", bold=True, size=18, color=DARK_BLUE)
for i, (fml, fmt) in enumerate(kpi_formulas):
    col_s = 2 + i*2
    ws1.merge_cells(start_row=4, start_column=col_s, end_row=4, end_column=col_s+1)
    sc(ws1, 4, col_s, fml, font=big_font, fill=PatternFill("solid", fgColor=LIGHT_BLUE), align=Alignment(horizontal="center"), fmt=fmt)

# Grade Distribution Table
section_label(ws1, "Grade Distribution", 6, 1, 5)
write_hdr(ws1, 7, 2, ["Grade", "Count", "Percentage", "Avg GPA"])
for idx, (gl, gn) in enumerate([("A",0),("B",1),("C",2),("D",3),("F",4)]):
    r = 8 + idx
    sc(ws1, r, 2, gl, font=bold_font, align=Alignment(horizontal="center"))
    sc(ws1, r, 3, '=COUNTIF(Data!O2:O' + str(DER) + ',' + str(gn) + ')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws1, r, 4, '=C' + str(r) + '/SUM($C$8:$C$12)', fmt=pct_fmt, align=Alignment(horizontal="center"))
    sc(ws1, r, 5, '=AVERAGEIF(Data!O2:O' + str(DER) + ',' + str(gn) + ',Data!N2:N' + str(DER) + ')', fmt=dec_fmt, align=Alignment(horizontal="center"))
sc(ws1, 13, 2, "Total", font=bold_font, align=Alignment(horizontal="center"))
sc(ws1, 13, 3, '=SUM(C8:C12)', font=bold_font, fmt=num_fmt, align=Alignment(horizontal="center"))
sc(ws1, 13, 4, '=SUM(D8:D12)', font=bold_font, fmt=pct_fmt, align=Alignment(horizontal="center"))
sc(ws1, 13, 5, '=AVERAGE(Data!N2:N' + str(DER) + ')', font=bold_font, fmt=dec_fmt, align=Alignment(horizontal="center"))
stripe(ws1, 8, 12, 2, 5)

# Additional Academic Stats (Formula-Driven)
section_label(ws1, "Key Academic Statistics (Formula-Driven)", 28, 1, 5)
write_hdr(ws1, 29, 2, ["Metric", "Value", "Formula"])
acad_stats = [
    ("Median GPA", '=MEDIAN(Data!N2:N'+str(DER)+')', "MEDIAN", dec_fmt),
    ("Std Dev GPA", '=ROUND(STDEV(Data!N2:N'+str(DER)+'),3)', "STDEV", '0.000'),
    ("Max GPA", '=MAX(Data!N2:N'+str(DER)+')', "MAX", dec_fmt),
    ("Min GPA", '=MIN(Data!N2:N'+str(DER)+')', "MIN", dec_fmt),
    ("GPA > 3.0 Count", '=COUNTIF(Data!N2:N'+str(DER)+',">3")', "COUNTIF", num_fmt),
    ("GPA > 3.0 Pct", '=COUNTIF(Data!N2:N'+str(DER)+',">3")/COUNTA(Data!A2:A'+str(DER)+')', "COUNTIF/COUNTA", pct_fmt),
    ("GPA < 2.0 Count", '=COUNTIF(Data!N2:N'+str(DER)+',"<2")', "COUNTIF", num_fmt),
    ("GPA < 2.0 Pct", '=COUNTIF(Data!N2:N'+str(DER)+',"<2")/COUNTA(Data!A2:A'+str(DER)+')', "COUNTIF/COUNTA", pct_fmt),
    ("Median Absences", '=MEDIAN(Data!G2:G'+str(DER)+')', "MEDIAN", '0.0'),
    ("Max Absences", '=MAX(Data!G2:G'+str(DER)+')', "MAX", num_fmt),
    ("Median Study Time", '=MEDIAN(Data!F2:F'+str(DER)+')', "MEDIAN", '0.0'),
    ("Max Study Time", '=MAX(Data!F2:F'+str(DER)+')', "MAX", '0.0'),
]
for idx, (mlbl, mfml, mexp, mfmt) in enumerate(acad_stats):
    r = 30 + idx
    sc(ws1, r, 2, mlbl, font=bold_font, align=Alignment(horizontal="left"))
    sc(ws1, r, 3, mfml, fmt=mfmt, align=Alignment(horizontal="center"))
    sc(ws1, r, 4, mexp, font=Font(name="Calibri", italic=True, size=9, color="666666"), align=Alignment(horizontal="left"))
stripe(ws1, 30, 41, 2, 4)

# Gender Breakdown (Formula)
section_label(ws1, "Gender Breakdown (Formula-Driven)", 43, 1, 5)
write_hdr(ws1, 44, 2, ["Gender", "Count", "Avg GPA", "Fail Rate"])
for idx, (glbl, gv) in enumerate([("Male",1),("Female",0)]):
    r = 45 + idx
    sc(ws1, r, 2, glbl, align=Alignment(horizontal="center"))
    sc(ws1, r, 3, '=COUNTIF(Data!C2:C'+str(DER)+','+str(gv)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws1, r, 4, '=AVERAGEIF(Data!C2:C'+str(DER)+','+str(gv)+',Data!N2:N'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws1, r, 5, '=COUNTIFS(Data!C2:C'+str(DER)+','+str(gv)+',Data!O2:O'+str(DER)+',4)/C'+str(r), fmt=pct_fmt, align=Alignment(horizontal="center"))
sc(ws1, 47, 2, "Total", font=bold_font, align=Alignment(horizontal="center"))
sc(ws1, 47, 3, '=SUM(C45:C46)', font=bold_font, fmt=num_fmt, align=Alignment(horizontal="center"))
sc(ws1, 47, 4, '=AVERAGE(Data!N2:N'+str(DER)+')', font=bold_font, fmt=dec_fmt, align=Alignment(horizontal="center"))
sc(ws1, 47, 5, '=COUNTIF(Data!O2:O'+str(DER)+',4)/COUNTA(Data!A2:A'+str(DER)+')', font=bold_font, fmt=pct_fmt, align=Alignment(horizontal="center"))
stripe(ws1, 45, 46, 2, 5)

# Extracurricular Activities Impact (Formula)
section_label(ws1, "Extracurricular Activities Impact (Formula-Driven)", 49, 1, 5)
write_hdr(ws1, 50, 2, ["Activity", "Participants", "Non-Participants Avg GPA", "Participants Avg GPA", "GPA Difference"])
activities = [("Extracurricular","J"), ("Sports","K"), ("Music","L"), ("Volunteering","M")]
for idx, (aname, acol) in enumerate(activities):
    r = 51 + idx
    sc(ws1, r, 2, aname, font=bold_font, align=Alignment(horizontal="left"))
    sc(ws1, r, 3, '=COUNTIF(Data!'+acol+'2:'+acol+str(DER)+',1)', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws1, r, 4, '=AVERAGEIF(Data!'+acol+'2:'+acol+str(DER)+',0,Data!N2:N'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws1, r, 5, '=AVERAGEIF(Data!'+acol+'2:'+acol+str(DER)+',1,Data!N2:N'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws1, r, 6, '=E'+str(r)+'-D'+str(r), fmt='+0.00;-0.00;0.00', align=Alignment(horizontal="center"))
stripe(ws1, 51, 54, 2, 6)

# CHART 1: Grade Distribution Bar
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Grade Distribution"
chart1.y_axis.title = "Number of Students"
chart1.x_axis.title = "Grade"
cats1 = Reference(ws1, min_col=2, min_row=8, max_row=12)
vals1 = Reference(ws1, min_col=3, min_row=7, max_row=12)
chart1.add_data(vals1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.width = 16
chart1.height = 10
chart1.series[0].graphicalProperties.solidFill = MED_BLUE
chart1.legend = None
dl1 = DataLabelList(); dl1.showVal = True
chart1.series[0].dLbls = dl1
ws1.add_chart(chart1, "G6")

# Pass vs Fail
section_label(ws1, "Pass vs Fail Breakdown", 15, 1, 5)
write_hdr(ws1, 16, 2, ["Status", "Count", "Percentage"])
sc(ws1, 17, 2, "Pass (A-D)", font=bold_font, align=Alignment(horizontal="center"))
sc(ws1, 17, 3, '=COUNTIF(Data!O2:O' + str(DER) + ',"<4")', fmt=num_fmt, align=Alignment(horizontal="center"))
sc(ws1, 17, 4, '=C17/SUM(C17:C18)', fmt=pct_fmt, align=Alignment(horizontal="center"))
sc(ws1, 18, 2, "Fail (F)", font=bold_font, align=Alignment(horizontal="center"))
sc(ws1, 18, 3, '=COUNTIF(Data!O2:O' + str(DER) + ',"=4")', fmt=num_fmt, align=Alignment(horizontal="center"))
sc(ws1, 18, 4, '=C18/SUM(C17:C18)', fmt=pct_fmt, align=Alignment(horizontal="center"))

# CHART 2: Pie Pass vs Fail
pie1 = PieChart()
pie1.title = "Pass vs Fail Rate"
pie1.style = 10
pie1.width = 14
pie1.height = 10
cats_p = Reference(ws1, min_col=2, min_row=17, max_row=18)
vals_p = Reference(ws1, min_col=3, min_row=16, max_row=18)
pie1.add_data(vals_p, titles_from_data=True)
pie1.set_categories(cats_p)
dp_pass = DataPoint(idx=0); dp_pass.graphicalProperties.solidFill = GREEN_FILL
dp_fail = DataPoint(idx=1); dp_fail.graphicalProperties.solidFill = RED_FILL
pie1.series[0].data_points = [dp_pass, dp_fail]
dl_p = DataLabelList(); dl_p.showPercent = True; dl_p.showCatName = True
pie1.series[0].dLbls = dl_p
ws1.add_chart(pie1, "G20")

# GPA by Parental Education
section_label(ws1, "Average GPA by Parental Education", 20, 1, 5)
write_hdr(ws1, 21, 2, ["Education Level", "Count", "Avg GPA", "Fail Rate"])
edu_levels = [(0,"None"),(1,"High School"),(2,"Some College"),(3,"Bachelor's"),(4,"Higher")]
for idx, (ev, el) in enumerate(edu_levels):
    r = 22 + idx
    sc(ws1, r, 2, el, align=Alignment(horizontal="center"))
    sc(ws1, r, 3, '=COUNTIF(Data!E2:E' + str(DER) + ',' + str(ev) + ')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws1, r, 4, '=AVERAGEIF(Data!E2:E' + str(DER) + ',' + str(ev) + ',Data!N2:N' + str(DER) + ')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws1, r, 5, '=COUNTIFS(Data!E2:E' + str(DER) + ',' + str(ev) + ',Data!O2:O' + str(DER) + ',4)/C' + str(r), fmt=pct_fmt, align=Alignment(horizontal="center"))
stripe(ws1, 22, 26, 2, 5)

# CHART 3: GPA by Education
chart_edu = BarChart()
chart_edu.type = "col"
chart_edu.style = 10
chart_edu.title = "Avg GPA by Parental Education"
chart_edu.y_axis.title = "GPA"
cats_edu = Reference(ws1, min_col=2, min_row=22, max_row=26)
vals_edu = Reference(ws1, min_col=4, min_row=21, max_row=26)
chart_edu.add_data(vals_edu, titles_from_data=True)
chart_edu.set_categories(cats_edu)
chart_edu.width = 16
chart_edu.height = 10
chart_edu.series[0].graphicalProperties.solidFill = ACCENT_GOLD
chart_edu.legend = None
dl_edu = DataLabelList(); dl_edu.showVal = True; dl_edu.numFmt = '0.00'
chart_edu.series[0].dLbls = dl_edu
ws1.add_chart(chart_edu, "G34")

for c in range(1, 20):
    ws1.column_dimensions[get_column_letter(c)].width = 14

# ============================================================
# PAGE 2: RISK FACTOR ANALYSIS
# ============================================================
ws2 = wb.create_sheet("2. Risk Factor Analysis")
sheet_title(ws2, "Risk Factor Analysis", row=1, merge_to=12)

# Study Time vs GPA
section_label(ws2, "Study Time vs Academic Performance", 3, 1, 7)
write_hdr(ws2, 4, 2, ["Study Time Band", "Students", "Avg GPA", "Fail Count", "Fail Rate"])
bands_st = [("0-5 hrs",0,5),("5-10 hrs",5,10),("10-15 hrs",10,15),("15-20 hrs",15,20),("20+ hrs",20,99)]
for idx, (lbl, lo, hi) in enumerate(bands_st):
    r = 5 + idx
    sc(ws2, r, 2, lbl, align=Alignment(horizontal="center"))
    if hi < 99:
        sc(ws2, r, 3, '=COUNTIFS(Data!F2:F'+str(DER)+',">="&'+str(lo)+',Data!F2:F'+str(DER)+',"<"&'+str(hi)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
        sc(ws2, r, 4, '=AVERAGEIFS(Data!N2:N'+str(DER)+',Data!F2:F'+str(DER)+',">="&'+str(lo)+',Data!F2:F'+str(DER)+',"<"&'+str(hi)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
        sc(ws2, r, 5, '=COUNTIFS(Data!F2:F'+str(DER)+',">="&'+str(lo)+',Data!F2:F'+str(DER)+',"<"&'+str(hi)+',Data!O2:O'+str(DER)+',4)', fmt=num_fmt, align=Alignment(horizontal="center"))
    else:
        sc(ws2, r, 3, '=COUNTIFS(Data!F2:F'+str(DER)+',">="&'+str(lo)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
        sc(ws2, r, 4, '=AVERAGEIFS(Data!N2:N'+str(DER)+',Data!F2:F'+str(DER)+',">="&'+str(lo)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
        sc(ws2, r, 5, '=COUNTIFS(Data!F2:F'+str(DER)+',">="&'+str(lo)+',Data!O2:O'+str(DER)+',4)', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 6, '=IF(C'+str(r)+'=0,0,E'+str(r)+'/C'+str(r)+')', fmt=pct_fmt, align=Alignment(horizontal="center"))
stripe(ws2, 5, 9, 2, 6)

# CHART: Study Time vs GPA bar
ch_st = BarChart()
ch_st.type = "col"; ch_st.style = 10
ch_st.title = "Average GPA by Study Time"
ch_st.y_axis.title = "GPA"; ch_st.x_axis.title = "Weekly Study Hours"
cats_st = Reference(ws2, min_col=2, min_row=5, max_row=9)
vals_st = Reference(ws2, min_col=4, min_row=4, max_row=9)
ch_st.add_data(vals_st, titles_from_data=True)
ch_st.set_categories(cats_st)
ch_st.width = 16; ch_st.height = 10
ch_st.series[0].graphicalProperties.solidFill = "4472C4"
ch_st.legend = None
dl_st = DataLabelList(); dl_st.showVal = True; dl_st.numFmt = '0.00'
ch_st.series[0].dLbls = dl_st
ws2.add_chart(ch_st, "H3")

# Absences vs GPA
section_label(ws2, "Absences vs Academic Performance", 12, 1, 7)
write_hdr(ws2, 13, 2, ["Absence Band", "Students", "Avg GPA", "Fail Count", "Fail Rate"])
bands_ab = [("0-5",0,5),("6-10",6,10),("11-15",11,15),("16-20",16,20),("21-25",21,25),("26-30",26,30)]
for idx, (lbl, lo, hi) in enumerate(bands_ab):
    r = 14 + idx
    sc(ws2, r, 2, lbl, align=Alignment(horizontal="center"))
    sc(ws2, r, 3, '=COUNTIFS(Data!G2:G'+str(DER)+',">="&'+str(lo)+',Data!G2:G'+str(DER)+',"<="&'+str(hi)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 4, '=AVERAGEIFS(Data!N2:N'+str(DER)+',Data!G2:G'+str(DER)+',">="&'+str(lo)+',Data!G2:G'+str(DER)+',"<="&'+str(hi)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 5, '=COUNTIFS(Data!G2:G'+str(DER)+',">="&'+str(lo)+',Data!G2:G'+str(DER)+',"<="&'+str(hi)+',Data!O2:O'+str(DER)+',4)', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 6, '=IF(C'+str(r)+'=0,0,E'+str(r)+'/C'+str(r)+')', fmt=pct_fmt, align=Alignment(horizontal="center"))
stripe(ws2, 14, 19, 2, 6)

# CHART: Absences vs GPA bar
ch_ab = BarChart()
ch_ab.type = "col"; ch_ab.style = 10
ch_ab.title = "Average GPA by Absence Band"
ch_ab.y_axis.title = "GPA"; ch_ab.x_axis.title = "Number of Absences"
cats_ab = Reference(ws2, min_col=2, min_row=14, max_row=19)
vals_ab = Reference(ws2, min_col=4, min_row=13, max_row=19)
ch_ab.add_data(vals_ab, titles_from_data=True)
ch_ab.set_categories(cats_ab)
ch_ab.width = 16; ch_ab.height = 10
ch_ab.series[0].graphicalProperties.solidFill = "ED7D31"
ch_ab.legend = None
dl_ab = DataLabelList(); dl_ab.showVal = True; dl_ab.numFmt = '0.00'
ch_ab.series[0].dLbls = dl_ab
ws2.add_chart(ch_ab, "H13")

# Fail Rate LINE chart
ch_fr = LineChart()
ch_fr.title = "Fail Rate Trend by Absence Band"
ch_fr.style = 10
ch_fr.y_axis.title = "Fail Rate"
ch_fr.y_axis.numFmt = '0%'
ch_fr.x_axis.title = "Absences"
cats_fr = Reference(ws2, min_col=2, min_row=14, max_row=19)
vals_fr = Reference(ws2, min_col=6, min_row=13, max_row=19)
ch_fr.add_data(vals_fr, titles_from_data=True)
ch_fr.set_categories(cats_fr)
ch_fr.width = 16; ch_fr.height = 10
ch_fr.series[0].graphicalProperties.line.solidFill = RED_FILL
ch_fr.series[0].graphicalProperties.line.width = 25000
ch_fr.legend = None
ws2.add_chart(ch_fr, "H27")

# Key Factor Correlations
section_label(ws2, "Key Factor Correlations with GPA", 22, 1, 7)
write_hdr(ws2, 23, 2, ["Factor", "Correlation with GPA", "Impact Level"])
factors = [("Study Time Weekly","F"),("Absences","G"),("Parental Support","I"),("Parental Education","E"),("Tutoring","H")]
for idx, (fname, fcol) in enumerate(factors):
    r = 24 + idx
    sc(ws2, r, 2, fname, align=Alignment(horizontal="center"))
    sc(ws2, r, 3, '=ROUND(CORREL(Data!'+fcol+'2:'+fcol+str(DER)+',Data!N2:N'+str(DER)+'),4)', fmt='0.0000', align=Alignment(horizontal="center"))
    sc(ws2, r, 4, '=IF(ABS(C'+str(r)+')>=0.5,"Strong",IF(ABS(C'+str(r)+')>=0.3,"Moderate",IF(ABS(C'+str(r)+')>=0.1,"Weak","Negligible")))', align=Alignment(horizontal="center"))
stripe(ws2, 24, 28, 2, 4)

# Tutoring Impact
section_label(ws2, "Tutoring Impact Analysis", 30, 1, 7)
write_hdr(ws2, 31, 2, ["Tutoring", "Students", "Avg GPA", "Fail Rate"])
for idx, (lbl, tv) in enumerate([("No Tutoring",0),("With Tutoring",1)]):
    r = 32 + idx
    sc(ws2, r, 2, lbl, align=Alignment(horizontal="center"))
    sc(ws2, r, 3, '=COUNTIF(Data!H2:H'+str(DER)+','+str(tv)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 4, '=AVERAGEIF(Data!H2:H'+str(DER)+','+str(tv)+',Data!N2:N'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 5, '=COUNTIFS(Data!H2:H'+str(DER)+','+str(tv)+',Data!O2:O'+str(DER)+',4)/C'+str(r), fmt=pct_fmt, align=Alignment(horizontal="center"))

# CHART: Tutoring
ch_tut = BarChart()
ch_tut.type = "col"; ch_tut.style = 10
ch_tut.title = "GPA: Tutoring vs No Tutoring"
ch_tut.y_axis.title = "Avg GPA"
cats_tut = Reference(ws2, min_col=2, min_row=32, max_row=33)
vals_tut = Reference(ws2, min_col=4, min_row=31, max_row=33)
ch_tut.add_data(vals_tut, titles_from_data=True)
ch_tut.set_categories(cats_tut)
ch_tut.width = 14; ch_tut.height = 10
ch_tut.series[0].graphicalProperties.solidFill = "548235"
ch_tut.legend = None
dl_tut = DataLabelList(); dl_tut.showVal = True; dl_tut.numFmt = '0.00'
ch_tut.series[0].dLbls = dl_tut
ws2.add_chart(ch_tut, "H41")

# Parental Support
section_label(ws2, "Parental Support Impact", 35, 1, 7)
write_hdr(ws2, 36, 2, ["Support Level", "Students", "Avg GPA", "Fail Rate", "Avg Risk Score"])
for idx, (sv, sl) in enumerate([(0,"None"),(1,"Low"),(2,"Moderate"),(3,"High"),(4,"Very High")]):
    r = 37 + idx
    sc(ws2, r, 2, sl, align=Alignment(horizontal="center"))
    sc(ws2, r, 3, '=COUNTIF(Data!I2:I'+str(DER)+','+str(sv)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 4, '=AVERAGEIF(Data!I2:I'+str(DER)+','+str(sv)+',Data!N2:N'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 5, '=COUNTIFS(Data!I2:I'+str(DER)+','+str(sv)+',Data!O2:O'+str(DER)+',4)/C'+str(r), fmt=pct_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 6, '=AVERAGEIF(Data!I2:I'+str(DER)+','+str(sv)+',Data!Q2:Q'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
stripe(ws2, 37, 41, 2, 5)

# Extracurricular Impact on GPA (Formula-Driven)
section_label(ws2, "Extracurricular Activities Impact (Formula-Driven)", 43, 1, 7)
write_hdr(ws2, 44, 2, ["Activity", "Participants", "Non-Part Count", "Part Avg GPA", "Non-Part Avg GPA", "Part Fail Rate", "Non-Part Fail Rate"])
act_factors = [("Extracurricular","J"), ("Sports","K"), ("Music","L"), ("Volunteering","M")]
for idx, (aname, acol) in enumerate(act_factors):
    r = 45 + idx
    sc(ws2, r, 2, aname, font=bold_font, align=Alignment(horizontal="left"))
    sc(ws2, r, 3, '=COUNTIF(Data!'+acol+'2:'+acol+str(DER)+',1)', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 4, '=COUNTIF(Data!'+acol+'2:'+acol+str(DER)+',0)', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 5, '=AVERAGEIF(Data!'+acol+'2:'+acol+str(DER)+',1,Data!N2:N'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 6, '=AVERAGEIF(Data!'+acol+'2:'+acol+str(DER)+',0,Data!N2:N'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 7, '=IF(C'+str(r)+'=0,0,COUNTIFS(Data!'+acol+'2:'+acol+str(DER)+',1,Data!O2:O'+str(DER)+',4)/C'+str(r)+')', fmt=pct_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 8, '=IF(D'+str(r)+'=0,0,COUNTIFS(Data!'+acol+'2:'+acol+str(DER)+',0,Data!O2:O'+str(DER)+',4)/D'+str(r)+')', fmt=pct_fmt, align=Alignment(horizontal="center"))
stripe(ws2, 45, 48, 2, 8)

# Age Group Analysis (Formula-Driven)
section_label(ws2, "Age Group Analysis (Formula-Driven)", 50, 1, 7)
write_hdr(ws2, 51, 2, ["Age Group", "Count", "Avg GPA", "Avg Absences", "Fail Rate"])
age_bands = [("15-16",15,16),("17-18",17,18)]
for idx, (albl, lo, hi) in enumerate(age_bands):
    r = 52 + idx
    sc(ws2, r, 2, albl, align=Alignment(horizontal="center"))
    sc(ws2, r, 3, '=COUNTIFS(Data!B2:B'+str(DER)+',">="&'+str(lo)+',Data!B2:B'+str(DER)+',"<="&'+str(hi)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 4, '=AVERAGEIFS(Data!N2:N'+str(DER)+',Data!B2:B'+str(DER)+',">="&'+str(lo)+',Data!B2:B'+str(DER)+',"<="&'+str(hi)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws2, r, 5, '=AVERAGEIFS(Data!G2:G'+str(DER)+',Data!B2:B'+str(DER)+',">="&'+str(lo)+',Data!B2:B'+str(DER)+',"<="&'+str(hi)+')', fmt='0.0', align=Alignment(horizontal="center"))
    sc(ws2, r, 6, '=IF(C'+str(r)+'=0,0,COUNTIFS(Data!B2:B'+str(DER)+',">="&'+str(lo)+',Data!B2:B'+str(DER)+',"<="&'+str(hi)+',Data!O2:O'+str(DER)+',4)/C'+str(r)+')', fmt=pct_fmt, align=Alignment(horizontal="center"))
stripe(ws2, 52, 53, 2, 6)

# Summary Statistics (Formula-Driven)
section_label(ws2, "Risk Factor Summary Statistics", 55, 1, 7)
write_hdr(ws2, 56, 2, ["Statistic", "Study Time", "Absences", "GPA", "Risk Score"])
summ_stats = [
    ("Mean", "AVERAGE"),
    ("Median", "MEDIAN"),
    ("Std Dev", "STDEV"),
    ("Min", "MIN"),
    ("Max", "MAX"),
]
stat_cols = {"Study Time":"F", "Absences":"G", "GPA":"N", "Risk Score":"Q"}
for idx, (slbl, sfunc) in enumerate(summ_stats):
    r = 57 + idx
    sc(ws2, r, 2, slbl, font=bold_font, align=Alignment(horizontal="left"))
    for ci, (_, dcol) in enumerate(stat_cols.items()):
        fmt_s = dec_fmt if dcol in ["N","Q"] else '0.0'
        sc(ws2, r, 3+ci, '=ROUND('+sfunc+'(Data!'+dcol+'2:'+dcol+str(DER)+'),2)', fmt=fmt_s, align=Alignment(horizontal="center"))
stripe(ws2, 57, 61, 2, 6)

for c in range(1, 20):
    ws2.column_dimensions[get_column_letter(c)].width = 15

# ============================================================
# PAGE 3: PERFORMANCE RISK INDEX
# ============================================================
ws3 = wb.create_sheet("3. Performance Risk Index")
sheet_title(ws3, "Performance Risk Index", row=1, merge_to=14)

# Risk Formula
section_label(ws3, "Risk Score Formula", 3, 1, 10)
ws3.merge_cells("B4:L4")
ws3.cell(row=4, column=2, value="Risk = (1-GPA/4)*35 + (Absences/Max)*25 + (1-StudyTime/Max)*20 + (1-Support/4)*10 + (Grade/4)*10")
ws3.cell(row=4, column=2).font = Font(name="Calibri", italic=True, size=10, color=MED_BLUE)
ws3.cell(row=4, column=2).alignment = Alignment(wrap_text=True)

# Risk Distribution
section_label(ws3, "Risk Category Distribution", 6, 1, 7)
write_hdr(ws3, 7, 2, ["Risk Category", "Score Range", "Count", "Percentage", "Avg GPA", "Avg Absences"])
risk_cats = [("Low","0-30",0,30),("Medium","30-55",30,55),("High","55-75",55,75),("Critical","75-100",75,100)]
risk_fills = [GREEN_FILL, YELLOW_FILL, ORANGE_FILL, RED_FILL]
for idx, (rl, rng, lo, hi) in enumerate(risk_cats):
    r = 8 + idx
    sc(ws3, r, 2, rl, font=bold_font, fill=PatternFill("solid", fgColor=risk_fills[idx]), align=Alignment(horizontal="center"))
    sc(ws3, r, 3, rng, align=Alignment(horizontal="center"))
    op = '"<="' if hi == 100 else '"<"'
    sc(ws3, r, 4, '=COUNTIFS(Data!Q2:Q'+str(DER)+',">="&'+str(lo)+',Data!Q2:Q'+str(DER)+','+op+'&'+str(hi)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws3, r, 5, '=D'+str(r)+'/SUM($D$8:$D$11)', fmt=pct_fmt, align=Alignment(horizontal="center"))
    sc(ws3, r, 6, '=AVERAGEIFS(Data!N2:N'+str(DER)+',Data!Q2:Q'+str(DER)+',">="&'+str(lo)+',Data!Q2:Q'+str(DER)+','+op+'&'+str(hi)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws3, r, 7, '=AVERAGEIFS(Data!G2:G'+str(DER)+',Data!Q2:Q'+str(DER)+',">="&'+str(lo)+',Data!Q2:Q'+str(DER)+','+op+'&'+str(hi)+')', fmt='0.0', align=Alignment(horizontal="center"))
sc(ws3, 12, 2, "Total", font=bold_font, align=Alignment(horizontal="center"))
sc(ws3, 12, 4, '=SUM(D8:D11)', font=bold_font, fmt=num_fmt, align=Alignment(horizontal="center"))
sc(ws3, 12, 5, '=SUM(E8:E11)', font=bold_font, fmt=pct_fmt, align=Alignment(horizontal="center"))

# CHART: Risk Pie
pie_risk = PieChart()
pie_risk.title = "Student Risk Distribution"
pie_risk.style = 10; pie_risk.width = 16; pie_risk.height = 11
cats_risk = Reference(ws3, min_col=2, min_row=8, max_row=11)
vals_risk = Reference(ws3, min_col=4, min_row=7, max_row=11)
pie_risk.add_data(vals_risk, titles_from_data=True)
pie_risk.set_categories(cats_risk)
for ci, clr in enumerate([GREEN_FILL, YELLOW_FILL, ORANGE_FILL, RED_FILL]):
    dp = DataPoint(idx=ci); dp.graphicalProperties.solidFill = clr
    pie_risk.series[0].data_points.append(dp)
dl_risk = DataLabelList(); dl_risk.showPercent = True; dl_risk.showCatName = True
pie_risk.series[0].dLbls = dl_risk
ws3.add_chart(pie_risk, "I6")

# CHART: Risk Bar
ch_rbar = BarChart()
ch_rbar.type = "col"; ch_rbar.style = 10
ch_rbar.title = "Students per Risk Category"
ch_rbar.y_axis.title = "Count"
cats_rb = Reference(ws3, min_col=2, min_row=8, max_row=11)
vals_rb = Reference(ws3, min_col=4, min_row=7, max_row=11)
ch_rbar.add_data(vals_rb, titles_from_data=True)
ch_rbar.set_categories(cats_rb)
ch_rbar.width = 16; ch_rbar.height = 11
for ci, clr in enumerate([GREEN_FILL, YELLOW_FILL, ORANGE_FILL, RED_FILL]):
    dp = DataPoint(idx=ci); dp.graphicalProperties.solidFill = clr
    ch_rbar.series[0].data_points.append(dp)
ch_rbar.legend = None
dl_rb = DataLabelList(); dl_rb.showVal = True
ch_rbar.series[0].dLbls = dl_rb
ws3.add_chart(ch_rbar, "I21")

# Top 25 At-Risk
section_label(ws3, "Top 25 At-Risk Students (Highest Risk Scores)", 14, 1, 9)
write_hdr(ws3, 15, 2, ["Rank", "Student ID", "GPA", "Grade", "Absences", "Study Hrs", "Risk Score", "Risk Level"])
top25 = df.nlargest(25, "RiskScore")
for idx, (_, stu) in enumerate(top25.iterrows()):
    r = 16 + idx
    sc(ws3, r, 2, idx+1, align=Alignment(horizontal="center"))
    sc(ws3, r, 3, int(stu["StudentID"]), align=Alignment(horizontal="center"))
    sc(ws3, r, 4, round(float(stu["GPA"]),2), fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws3, r, 5, str(stu["GradeLetter"]), align=Alignment(horizontal="center"))
    sc(ws3, r, 6, int(stu["Absences"]), align=Alignment(horizontal="center"))
    sc(ws3, r, 7, round(float(stu["StudyTimeWeekly"]),1), fmt='0.0', align=Alignment(horizontal="center"))
    sc(ws3, r, 8, round(float(stu["RiskScore"]),1), fmt='0.0', align=Alignment(horizontal="center"))
    rcat = str(stu["RiskCategory"])
    rf = RED_FILL if rcat == "Critical" else ORANGE_FILL
    sc(ws3, r, 9, rcat, font=bold_font, fill=PatternFill("solid", fgColor=rf), align=Alignment(horizontal="center"))
stripe(ws3, 16, 40, 2, 9)

# Color scale on risk scores
ws3.conditional_formatting.add('H16:H40',
    ColorScaleRule(start_type='num', start_value=50, start_color='70AD47',
                   mid_type='num', mid_value=75, mid_color='FFD966',
                   end_type='num', end_value=100, end_color='FF4444'))

# Risk Score Summary Statistics (Formula-Driven)
section_label(ws3, "Risk Score Statistics (Formula-Driven)", 42, 1, 9)
write_hdr(ws3, 43, 2, ["Statistic", "Value", "Formula Used"])
risk_stats = [
    ("Average Risk Score", '=ROUND(AVERAGE(Data!Q2:Q'+str(DER)+'),2)', "AVERAGE", dec_fmt),
    ("Median Risk Score", '=ROUND(MEDIAN(Data!Q2:Q'+str(DER)+'),2)', "MEDIAN", dec_fmt),
    ("Std Dev Risk Score", '=ROUND(STDEV(Data!Q2:Q'+str(DER)+'),3)', "STDEV", '0.000'),
    ("Max Risk Score", '=ROUND(MAX(Data!Q2:Q'+str(DER)+'),2)', "MAX", dec_fmt),
    ("Min Risk Score", '=ROUND(MIN(Data!Q2:Q'+str(DER)+'),2)', "MIN", dec_fmt),
    ("Students Risk > 75", '=COUNTIF(Data!Q2:Q'+str(DER)+',">75")', "COUNTIF", num_fmt),
    ("Students Risk > 55", '=COUNTIF(Data!Q2:Q'+str(DER)+',">55")', "COUNTIF", num_fmt),
    ("Students Risk < 30", '=COUNTIF(Data!Q2:Q'+str(DER)+',"<30")', "COUNTIF", num_fmt),
    ("Pct Critical+High", '=(COUNTIF(Data!Q2:Q'+str(DER)+',">55"))/COUNTA(Data!A2:A'+str(DER)+')', "COUNTIF/COUNTA", pct_fmt),
    ("Pct Low Risk", '=COUNTIF(Data!Q2:Q'+str(DER)+',"<30")/COUNTA(Data!A2:A'+str(DER)+')', "COUNTIF/COUNTA", pct_fmt),
]
for idx, (rlbl, rfml, rexp, rfmt) in enumerate(risk_stats):
    r = 44 + idx
    sc(ws3, r, 2, rlbl, font=bold_font, align=Alignment(horizontal="left"))
    sc(ws3, r, 3, rfml, fmt=rfmt, align=Alignment(horizontal="center"))
    sc(ws3, r, 4, rexp, font=Font(name="Calibri", italic=True, size=9, color="666666"), align=Alignment(horizontal="left"))
stripe(ws3, 44, 53, 2, 4)

# Risk by Grade (Formula-Driven Cross-Tab)
section_label(ws3, "Risk Distribution by Grade (Formula-Driven)", 55, 1, 9)
write_hdr(ws3, 56, 2, ["Grade", "Count", "Avg Risk Score", "High+Critical Count", "High+Critical %"])
for idx, (gl, gn) in enumerate([("A",0),("B",1),("C",2),("D",3),("F",4)]):
    r = 57 + idx
    sc(ws3, r, 2, gl, font=bold_font, align=Alignment(horizontal="center"))
    sc(ws3, r, 3, '=COUNTIF(Data!O2:O'+str(DER)+','+str(gn)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws3, r, 4, '=AVERAGEIF(Data!O2:O'+str(DER)+','+str(gn)+',Data!Q2:Q'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws3, r, 5, '=COUNTIFS(Data!O2:O'+str(DER)+','+str(gn)+',Data!Q2:Q'+str(DER)+',">55")', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws3, r, 6, '=IF(C'+str(r)+'=0,0,E'+str(r)+'/C'+str(r)+')', fmt=pct_fmt, align=Alignment(horizontal="center"))
sc(ws3, 62, 2, "Total", font=bold_font, align=Alignment(horizontal="center"))
sc(ws3, 62, 3, '=SUM(C57:C61)', font=bold_font, fmt=num_fmt, align=Alignment(horizontal="center"))
sc(ws3, 62, 4, '=AVERAGE(Data!Q2:Q'+str(DER)+')', font=bold_font, fmt=dec_fmt, align=Alignment(horizontal="center"))
sc(ws3, 62, 5, '=SUM(E57:E61)', font=bold_font, fmt=num_fmt, align=Alignment(horizontal="center"))
sc(ws3, 62, 6, '=E62/C62', font=bold_font, fmt=pct_fmt, align=Alignment(horizontal="center"))
stripe(ws3, 57, 61, 2, 6)

# Risk by Tutoring Status (Formula)
section_label(ws3, "Risk by Tutoring Status (Formula-Driven)", 64, 1, 9)
write_hdr(ws3, 65, 2, ["Tutoring", "Students", "Avg Risk Score", "High+Critical Count", "Pct High Risk"])
for idx, (tlbl, tv) in enumerate([("No Tutoring",0),("With Tutoring",1)]):
    r = 66 + idx
    sc(ws3, r, 2, tlbl, align=Alignment(horizontal="center"))
    sc(ws3, r, 3, '=COUNTIF(Data!H2:H'+str(DER)+','+str(tv)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws3, r, 4, '=AVERAGEIF(Data!H2:H'+str(DER)+','+str(tv)+',Data!Q2:Q'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws3, r, 5, '=COUNTIFS(Data!H2:H'+str(DER)+','+str(tv)+',Data!Q2:Q'+str(DER)+',">55")', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws3, r, 6, '=IF(C'+str(r)+'=0,0,E'+str(r)+'/C'+str(r)+')', fmt=pct_fmt, align=Alignment(horizontal="center"))
stripe(ws3, 66, 67, 2, 6)

for c in range(1, 20):
    ws3.column_dimensions[get_column_letter(c)].width = 14

# ============================================================
# PAGE 4: INTERVENTION STRATEGY SIMULATOR
# ============================================================
ws4 = wb.create_sheet("4. Intervention Simulator")
sheet_title(ws4, "Intervention Strategy Simulator", row=1, merge_to=14)

# Scenario Params
section_label(ws4, "Scenario Parameters (Adjust Yellow Cells)", 3, 1, 8)
write_hdr(ws4, 4, 2, ["Parameter", "", "Value", "Unit"])
input_fill = PatternFill("solid", fgColor=YELLOW_FILL)
params = [
    ("Study Time Increase", 5, "hours/week"),
    ("Absence Reduction", 5, "fewer absences"),
    ("Tutoring Enrollment", 20, "% of at-risk"),
    ("Mentoring Program", 15, "% of at-risk"),
    ("Cost per Tutor/Semester", 500, "USD"),
    ("Cost per Mentor/Semester", 300, "USD"),
]
for idx, (lbl, val, unit) in enumerate(params):
    r = 5 + idx
    ws4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    sc(ws4, r, 2, lbl, font=bold_font, align=Alignment(horizontal="left"))
    sc(ws4, r, 4, val, fill=input_fill, font=Font(name="Calibri", bold=True, size=12), align=Alignment(horizontal="center"))
    sc(ws4, r, 5, unit, align=Alignment(horizontal="left"))

# Current State
section_label(ws4, "Current State (Formula-Driven)", 12, 1, 8)
write_hdr(ws4, 13, 2, ["Metric", "Value", "Formula Used"])
curr_metrics = [
    ("Total Students", '=COUNTA(Data!A2:A'+str(DER)+')', "COUNTA", num_fmt),
    ("Current Fail Count", '=COUNTIF(Data!O2:O'+str(DER)+',4)', "COUNTIF(Grade=4)", num_fmt),
    ("Current Fail Rate", '=C15/C14', "Fails/Total", pct_fmt),
    ("At-Risk Students (Score>55)", '=COUNTIF(Data!Q2:Q'+str(DER)+',">55")', "COUNTIF(Risk>55)", num_fmt),
    ("Avg GPA of At-Risk", '=AVERAGEIF(Data!Q2:Q'+str(DER)+',">55",Data!N2:N'+str(DER)+')', "AVERAGEIF", dec_fmt),
    ("Avg Absences of At-Risk", '=AVERAGEIF(Data!Q2:Q'+str(DER)+',">55",Data!G2:G'+str(DER)+')', "AVERAGEIF", '0.0'),
]
for idx, (mlbl, mfml, mexp, mfmt) in enumerate(curr_metrics):
    r = 14 + idx
    sc(ws4, r, 2, mlbl, font=bold_font, align=Alignment(horizontal="left"))
    sc(ws4, r, 3, mfml, fmt=mfmt, align=Alignment(horizontal="center"))
    sc(ws4, r, 4, mexp, font=Font(name="Calibri", italic=True, size=9, color="666666"), align=Alignment(horizontal="left"))
stripe(ws4, 14, 19, 2, 4)

# Projections
section_label(ws4, "Projected Improvement (Scenario-Based)", 21, 1, 8)
write_hdr(ws4, 22, 2, ["Projection", "Value", "Calculation"])
proj = [
    ("GPA Lift from Study Time", '=ROUND(D5*0.04,2)', "+0.04 GPA per hr", dec_fmt),
    ("GPA Lift from Fewer Absences", '=ROUND(D6*0.03,2)', "+0.03 GPA per absence", dec_fmt),
    ("Combined GPA Improvement", '=C23+C24', "Sum of lifts", dec_fmt),
    ("New At-Risk Avg GPA", '=C18+C25', "Current + improvement", dec_fmt),
    ("Est. Students Saved from Fail", '=ROUND(C15*(C25/(4-C18)),0)', "Proportional", num_fmt),
    ("New Projected Fail Rate", '=MAX(0,(C15-C27)/C14)', "Adjusted", pct_fmt),
    ("Fail Rate Reduction", '=C16-C28', "Improvement", pct_fmt),
    ("Target Met (20% reduction)?", '=IF(C29>=C16*0.2,"YES - Target Met","NO - Increase Parameters")', ">=20%", None),
]
for idx, (plbl, pfml, pexp, pfmt) in enumerate(proj):
    r = 23 + idx
    sc(ws4, r, 2, plbl, font=bold_font, align=Alignment(horizontal="left"))
    sc(ws4, r, 3, pfml, fmt=pfmt, align=Alignment(horizontal="center"))
    sc(ws4, r, 4, pexp, font=Font(name="Calibri", italic=True, size=9, color="666666"), align=Alignment(horizontal="left"))
stripe(ws4, 23, 30, 2, 4)

# Cost-Benefit
section_label(ws4, "Cost-Benefit Estimate", 32, 1, 8)
write_hdr(ws4, 33, 2, ["Cost Item", "Value", "Formula"])
costs = [
    ("Students Needing Tutoring", '=ROUND(C17*D7/100,0)', "At-Risk x %", num_fmt),
    ("Students Needing Mentoring", '=ROUND(C17*D8/100,0)', "At-Risk x %", num_fmt),
    ("Total Tutoring Cost", '=C34*D9', "Count x Rate", '"$"#,##0'),
    ("Total Mentoring Cost", '=C35*D10', "Count x Rate", '"$"#,##0'),
    ("Total Program Cost", '=C36+C37', "Sum", '"$"#,##0'),
    ("Cost per Student Saved", '=IF(C27>0,ROUND(C38/C27,0),0)', "Total/Saved", '"$"#,##0'),
]
for idx, (clbl, cfml, cexp, cfmt) in enumerate(costs):
    r = 34 + idx
    sc(ws4, r, 2, clbl, font=bold_font, align=Alignment(horizontal="left"))
    sc(ws4, r, 3, cfml, fmt=cfmt, align=Alignment(horizontal="center"))
    sc(ws4, r, 4, cexp, font=Font(name="Calibri", italic=True, size=9, color="666666"), align=Alignment(horizontal="left"))
stripe(ws4, 34, 39, 2, 4)

# Resource Allocation by Risk Level (Formula-Driven)
section_label(ws4, "Resource Allocation by Risk Level (Formula-Driven)", 40, 1, 8)
write_hdr(ws4, 41, 2, ["Risk Level", "Students", "% of Total", "Tutoring Alloc", "Mentoring Alloc", "Tutoring Cost", "Mentoring Cost"])
risk_alloc = [("Low",0,30,0,0), ("Medium",30,55,10,5), ("High",55,75,40,30), ("Critical",75,100,50,65)]
for idx, (rl, lo, hi, tpct, mpct) in enumerate(risk_alloc):
    r = 42 + idx
    op = '"<="' if hi == 100 else '"<"'
    sc(ws4, r, 2, rl, font=bold_font, align=Alignment(horizontal="center"))
    sc(ws4, r, 3, '=COUNTIFS(Data!Q2:Q'+str(DER)+',">="&'+str(lo)+',Data!Q2:Q'+str(DER)+','+op+'&'+str(hi)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws4, r, 4, '=C'+str(r)+'/COUNTA(Data!A2:A'+str(DER)+')', fmt=pct_fmt, align=Alignment(horizontal="center"))
    sc(ws4, r, 5, '=ROUND(C'+str(r)+'*'+str(tpct)+'/100,0)', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws4, r, 6, '=ROUND(C'+str(r)+'*'+str(mpct)+'/100,0)', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws4, r, 7, '=E'+str(r)+'*D9', fmt='"$"#,##0', align=Alignment(horizontal="center"))
    sc(ws4, r, 8, '=F'+str(r)+'*D10', fmt='"$"#,##0', align=Alignment(horizontal="center"))
sc(ws4, 46, 2, "Total", font=bold_font, align=Alignment(horizontal="center"))
sc(ws4, 46, 3, '=SUM(C42:C45)', font=bold_font, fmt=num_fmt, align=Alignment(horizontal="center"))
sc(ws4, 46, 4, '=SUM(D42:D45)', font=bold_font, fmt=pct_fmt, align=Alignment(horizontal="center"))
sc(ws4, 46, 5, '=SUM(E42:E45)', font=bold_font, fmt=num_fmt, align=Alignment(horizontal="center"))
sc(ws4, 46, 6, '=SUM(F42:F45)', font=bold_font, fmt=num_fmt, align=Alignment(horizontal="center"))
sc(ws4, 46, 7, '=SUM(G42:G45)', font=bold_font, fmt='"$"#,##0', align=Alignment(horizontal="center"))
sc(ws4, 46, 8, '=SUM(H42:H45)', font=bold_font, fmt='"$"#,##0', align=Alignment(horizontal="center"))
stripe(ws4, 42, 45, 2, 8)

# ROI Metrics (Formula-Driven)
section_label(ws4, "Return on Investment Metrics (Formula-Driven)", 48, 1, 8)
write_hdr(ws4, 49, 2, ["ROI Metric", "Value", "Formula"])
roi_metrics = [
    ("Total Intervention Budget", '=G46+H46', "Sum of costs", '"$"#,##0'),
    ("Students Targeted", '=E46+F46', "Sum tutoring + mentoring", num_fmt),
    ("Cost per At-Risk Student", '=IF((E46+F46)>0,ROUND((G46+H46)/(E46+F46),0),0)', "Budget/Targeted", '"$"#,##0'),
    ("Est. Students Saved", '=C27', "From projections", num_fmt),
    ("Cost per Student Saved", '=IF(C27>0,ROUND((G46+H46)/C27,0),0)', "Budget/Saved", '"$"#,##0'),
    ("Dropout Cost Avoided (est $30K/student)", '=C27*30000', "Saved x $30K", '"$"#,##0'),
    ("Net ROI", '=C27*30000-(G46+H46)', "Avoided - Budget", '"$"#,##0'),
]
for idx, (mlbl, mfml, mexp, mfmt) in enumerate(roi_metrics):
    r = 50 + idx
    sc(ws4, r, 2, mlbl, font=bold_font, align=Alignment(horizontal="left"))
    sc(ws4, r, 3, mfml, fmt=mfmt, align=Alignment(horizontal="center"))
    sc(ws4, r, 4, mexp, font=Font(name="Calibri", italic=True, size=9, color="666666"), align=Alignment(horizontal="left"))
stripe(ws4, 50, 56, 2, 4)

# CHART: Current vs Projected
section_label(ws4, "Current vs Projected", 58, 1, 6)
write_hdr(ws4, 59, 2, ["Metric", "Current", "Projected"])
sc(ws4, 60, 2, "Fail Rate", font=bold_font, align=Alignment(horizontal="center"))
sc(ws4, 60, 3, '=C16', fmt=pct_fmt, align=Alignment(horizontal="center"))
sc(ws4, 60, 4, '=C28', fmt=pct_fmt, align=Alignment(horizontal="center"))
sc(ws4, 61, 2, "Avg GPA (At-Risk)", font=bold_font, align=Alignment(horizontal="center"))
sc(ws4, 61, 3, '=C18', fmt=dec_fmt, align=Alignment(horizontal="center"))
sc(ws4, 61, 4, '=C26', fmt=dec_fmt, align=Alignment(horizontal="center"))
sc(ws4, 62, 2, "Fail Count", font=bold_font, align=Alignment(horizontal="center"))
sc(ws4, 62, 3, '=C15', fmt=num_fmt, align=Alignment(horizontal="center"))
sc(ws4, 62, 4, '=MAX(0,C15-C27)', fmt=num_fmt, align=Alignment(horizontal="center"))

ch_comp = BarChart()
ch_comp.type = "col"; ch_comp.grouping = "clustered"; ch_comp.style = 10
ch_comp.title = "Current vs Projected After Intervention"
ch_comp.y_axis.title = "Value"
cats_comp = Reference(ws4, min_col=2, min_row=60, max_row=62)
vals_curr = Reference(ws4, min_col=3, min_row=59, max_row=62)
vals_proj = Reference(ws4, min_col=4, min_row=59, max_row=62)
ch_comp.add_data(vals_curr, titles_from_data=True)
ch_comp.add_data(vals_proj, titles_from_data=True)
ch_comp.set_categories(cats_comp)
ch_comp.width = 18; ch_comp.height = 12
ch_comp.series[0].graphicalProperties.solidFill = RED_FILL
ch_comp.series[1].graphicalProperties.solidFill = GREEN_FILL
ws4.add_chart(ch_comp, "F12")

# Cost Pie
sc(ws4, 64, 6, "Tutoring", align=Alignment(horizontal="center"))
sc(ws4, 65, 6, "Mentoring", align=Alignment(horizontal="center"))
sc(ws4, 64, 7, '=C36', fmt='"$"#,##0', align=Alignment(horizontal="center"))
sc(ws4, 65, 7, '=C37', fmt='"$"#,##0', align=Alignment(horizontal="center"))
pie_cost = PieChart()
pie_cost.title = "Cost Breakdown"; pie_cost.style = 10
pie_cost.width = 14; pie_cost.height = 10
cats_cost = Reference(ws4, min_col=6, min_row=64, max_row=65)
vals_cost = Reference(ws4, min_col=7, min_row=64, max_row=65)
pie_cost.add_data(vals_cost, titles_from_data=False)
pie_cost.set_categories(cats_cost)
dp_c0 = DataPoint(idx=0); dp_c0.graphicalProperties.solidFill = "4472C4"
dp_c1 = DataPoint(idx=1); dp_c1.graphicalProperties.solidFill = ACCENT_GOLD
pie_cost.series[0].data_points = [dp_c0, dp_c1]
dl_cost = DataLabelList(); dl_cost.showPercent = True; dl_cost.showCatName = True
pie_cost.series[0].dLbls = dl_cost
ws4.add_chart(pie_cost, "F29")

for c in range(1, 20):
    ws4.column_dimensions[get_column_letter(c)].width = 18

# ============================================================
# PAGE 5: ETHICS & SAFEGUARDS
# ============================================================
ws5 = wb.create_sheet("5. Ethics & Safeguards")
sheet_title(ws5, "Ethics, Privacy & Safeguards", row=1, merge_to=12)

# Transparency Matrix
section_label(ws5, "Labeling Transparency Matrix", 3, 1, 8)
write_hdr(ws5, 4, 2, ["Factor", "Weight", "Why Used", "Bias Risk", "Mitigation"])
transparency = [
    ("GPA","35%","Primary academic indicator","Low - objective","Validated by registrar"),
    ("Absences","25%","Predictor of disengagement","Medium - health/family","Context review before action"),
    ("Study Time","20%","Effort indicator","Medium - self-reported","Cross-reference with grades"),
    ("Parental Support","10%","Environmental factor","High - socioeconomic","Never used alone"),
    ("Grade Class","10%","Current standing","Low - objective","Confirmation only"),
]
for idx, (f, w, why, bias, mit) in enumerate(transparency):
    r = 5 + idx
    sc(ws5, r, 2, f, font=bold_font, align=Alignment(horizontal="center"))
    sc(ws5, r, 3, w, align=Alignment(horizontal="center"))
    sc(ws5, r, 4, why, align=Alignment(horizontal="left", wrap_text=True))
    sc(ws5, r, 5, bias, align=Alignment(horizontal="left", wrap_text=True))
    sc(ws5, r, 6, mit, align=Alignment(horizontal="left", wrap_text=True))
stripe(ws5, 5, 9, 2, 6)

# Factor Weight chart data
write_hdr(ws5, 4, 8, ["Factor", "Weight %"])
wts = [("GPA",35),("Absences",25),("Study Time",20),("Parental Support",10),("Grade Class",10)]
for idx, (fn, fw) in enumerate(wts):
    sc(ws5, 5+idx, 8, fn, align=Alignment(horizontal="center"))
    sc(ws5, 5+idx, 9, fw, align=Alignment(horizontal="center"))
pie_wt = PieChart()
pie_wt.title = "Risk Score Factor Weights"
pie_wt.style = 10; pie_wt.width = 16; pie_wt.height = 11
cats_wt = Reference(ws5, min_col=8, min_row=5, max_row=9)
vals_wt = Reference(ws5, min_col=9, min_row=4, max_row=9)
pie_wt.add_data(vals_wt, titles_from_data=True)
pie_wt.set_categories(cats_wt)
for ci, clr in enumerate([MED_BLUE,"ED7D31","A5A5A5",ACCENT_GOLD,"4472C4"]):
    dp = DataPoint(idx=ci); dp.graphicalProperties.solidFill = clr
    pie_wt.series[0].data_points.append(dp)
dl_wt = DataLabelList(); dl_wt.showPercent = True; dl_wt.showCatName = True
pie_wt.series[0].dLbls = dl_wt
ws5.add_chart(pie_wt, "H11")

# Privacy Policy
section_label(ws5, "Student Data Privacy Policy", 12, 1, 7)
write_hdr(ws5, 13, 2, ["Policy Area", "Description"])
policies = [
    ("Data Minimization","Only academically relevant data used. No names or addresses in scoring."),
    ("Purpose Limitation","Data used solely for early intervention, not disciplinary decisions."),
    ("Access Control","Dashboard access restricted to authorized academic advisors only."),
    ("Consent & Transparency","Students informed about data usage. Opt-out available."),
    ("Retention Policy","Risk scores recalculated each semester. History purged after 2 years."),
    ("Right to Challenge","Students may contest risk categorization via academic affairs."),
    ("Bias Auditing","Annual review across demographics to detect disparate impact."),
    ("De-identification","All reports use student IDs only. Re-identification requires secure system."),
]
for idx, (area, desc) in enumerate(policies):
    r = 14 + idx
    sc(ws5, r, 2, area, font=bold_font, align=Alignment(horizontal="left", vertical="top"))
    ws5.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    sc(ws5, r, 3, desc, align=Alignment(horizontal="left", wrap_text=True, vertical="top"))
    ws5.row_dimensions[r].height = 30
stripe(ws5, 14, 21, 2, 6)

# Ethical Guidelines
section_label(ws5, "Ethical Usage Guidelines", 23, 1, 7)
write_hdr(ws5, 24, 2, ["Principle", "Guideline", "Implementation"])
guidelines = [
    ("Non-Punitive","Risk labels NEVER used for punishment","Training + audit trail"),
    ("Human-in-the-Loop","No automated decisions; advisor review required","System flags only"),
    ("Equity Focus","Monitor disparate impact across demographics","Quarterly bias reports"),
    ("Student Agency","Students can improve risk score through action","Clear guidance provided"),
    ("Continuous Improvement","Model reviewed each semester","Track outcomes; adjust weights"),
]
for idx, (pr, gd, impl) in enumerate(guidelines):
    r = 25 + idx
    sc(ws5, r, 2, pr, font=bold_font, align=Alignment(horizontal="left", vertical="top"))
    sc(ws5, r, 3, gd, align=Alignment(horizontal="left", wrap_text=True, vertical="top"))
    sc(ws5, r, 4, impl, align=Alignment(horizontal="left", wrap_text=True, vertical="top"))
    ws5.row_dimensions[r].height = 35
stripe(ws5, 25, 29, 2, 4)

# Demographic Fairness
section_label(ws5, "Demographic Fairness Check (Formula-Driven)", 31, 1, 8)
write_hdr(ws5, 32, 2, ["Group", "Count", "Avg GPA", "Fail Rate", "Avg Risk Score"])
for idx, (glbl, gv) in enumerate([("Male (Gender=1)",1),("Female (Gender=0)",0)]):
    r = 33 + idx
    sc(ws5, r, 2, glbl, align=Alignment(horizontal="center"))
    sc(ws5, r, 3, '=COUNTIF(Data!C2:C'+str(DER)+','+str(gv)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 4, '=AVERAGEIF(Data!C2:C'+str(DER)+','+str(gv)+',Data!N2:N'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 5, '=COUNTIFS(Data!C2:C'+str(DER)+','+str(gv)+',Data!O2:O'+str(DER)+',4)/C'+str(r), fmt=pct_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 6, '=AVERAGEIF(Data!C2:C'+str(DER)+','+str(gv)+',Data!Q2:Q'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
for idx, ev in enumerate(range(4)):
    r = 35 + idx
    sc(ws5, r, 2, "Ethnicity "+str(ev), align=Alignment(horizontal="center"))
    sc(ws5, r, 3, '=COUNTIF(Data!D2:D'+str(DER)+','+str(ev)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 4, '=AVERAGEIF(Data!D2:D'+str(DER)+','+str(ev)+',Data!N2:N'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 5, '=COUNTIFS(Data!D2:D'+str(DER)+','+str(ev)+',Data!O2:O'+str(DER)+',4)/C'+str(r), fmt=pct_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 6, '=AVERAGEIF(Data!D2:D'+str(DER)+','+str(ev)+',Data!Q2:Q'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
stripe(ws5, 33, 38, 2, 6)

# Ethnicity detailed analysis with more formulas
section_label(ws5, "Ethnicity Detailed Analysis (Formula-Driven)", 40, 1, 8)
write_hdr(ws5, 41, 2, ["Ethnicity", "Count", "Avg GPA", "Avg Study Time", "Avg Absences", "Fail Rate", "Avg Risk Score"])
for idx, ev in enumerate(range(4)):
    r = 42 + idx
    sc(ws5, r, 2, "Ethnicity "+str(ev), align=Alignment(horizontal="center"))
    sc(ws5, r, 3, '=COUNTIF(Data!D2:D'+str(DER)+','+str(ev)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 4, '=AVERAGEIF(Data!D2:D'+str(DER)+','+str(ev)+',Data!N2:N'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 5, '=AVERAGEIF(Data!D2:D'+str(DER)+','+str(ev)+',Data!F2:F'+str(DER)+')', fmt='0.0', align=Alignment(horizontal="center"))
    sc(ws5, r, 6, '=AVERAGEIF(Data!D2:D'+str(DER)+','+str(ev)+',Data!G2:G'+str(DER)+')', fmt='0.0', align=Alignment(horizontal="center"))
    sc(ws5, r, 7, '=IF(C'+str(r)+'=0,0,COUNTIFS(Data!D2:D'+str(DER)+','+str(ev)+',Data!O2:O'+str(DER)+',4)/C'+str(r)+')', fmt=pct_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 8, '=AVERAGEIF(Data!D2:D'+str(DER)+','+str(ev)+',Data!Q2:Q'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
sc(ws5, 46, 2, "Overall", font=bold_font, align=Alignment(horizontal="center"))
sc(ws5, 46, 3, '=COUNTA(Data!A2:A'+str(DER)+')', font=bold_font, fmt=num_fmt, align=Alignment(horizontal="center"))
sc(ws5, 46, 4, '=AVERAGE(Data!N2:N'+str(DER)+')', font=bold_font, fmt=dec_fmt, align=Alignment(horizontal="center"))
sc(ws5, 46, 5, '=AVERAGE(Data!F2:F'+str(DER)+')', font=bold_font, fmt='0.0', align=Alignment(horizontal="center"))
sc(ws5, 46, 6, '=AVERAGE(Data!G2:G'+str(DER)+')', font=bold_font, fmt='0.0', align=Alignment(horizontal="center"))
sc(ws5, 46, 7, '=COUNTIF(Data!O2:O'+str(DER)+',4)/COUNTA(Data!A2:A'+str(DER)+')', font=bold_font, fmt=pct_fmt, align=Alignment(horizontal="center"))
sc(ws5, 46, 8, '=AVERAGE(Data!Q2:Q'+str(DER)+')', font=bold_font, fmt=dec_fmt, align=Alignment(horizontal="center"))
stripe(ws5, 42, 45, 2, 8)

# Parental Education Fairness (Formula-Driven)
section_label(ws5, "Parental Education Equity Check (Formula-Driven)", 48, 1, 8)
write_hdr(ws5, 49, 2, ["Education Level", "Count", "Avg GPA", "Fail Rate", "Avg Risk Score", "High Risk %"])
edu_levels_eth = [(0,"None"),(1,"High School"),(2,"Some College"),(3,"Bachelor's"),(4,"Higher")]
for idx, (ev, el) in enumerate(edu_levels_eth):
    r = 50 + idx
    sc(ws5, r, 2, el, align=Alignment(horizontal="center"))
    sc(ws5, r, 3, '=COUNTIF(Data!E2:E'+str(DER)+','+str(ev)+')', fmt=num_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 4, '=AVERAGEIF(Data!E2:E'+str(DER)+','+str(ev)+',Data!N2:N'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 5, '=IF(C'+str(r)+'=0,0,COUNTIFS(Data!E2:E'+str(DER)+','+str(ev)+',Data!O2:O'+str(DER)+',4)/C'+str(r)+')', fmt=pct_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 6, '=AVERAGEIF(Data!E2:E'+str(DER)+','+str(ev)+',Data!Q2:Q'+str(DER)+')', fmt=dec_fmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 7, '=IF(C'+str(r)+'=0,0,COUNTIFS(Data!E2:E'+str(DER)+','+str(ev)+',Data!Q2:Q'+str(DER)+',">55")/C'+str(r)+')', fmt=pct_fmt, align=Alignment(horizontal="center"))
stripe(ws5, 50, 54, 2, 7)

# Disparity Analysis (Formula-Driven)
section_label(ws5, "Disparity Analysis (Formula-Driven)", 56, 1, 8)
write_hdr(ws5, 57, 2, ["Metric", "Value", "Formula Used"])
disparity_metrics = [
    ("Max Fail Rate (Gender)", '=MAX(E33:E34)', "MAX", pct_fmt),
    ("Min Fail Rate (Gender)", '=MIN(E33:E34)', "MIN", pct_fmt),
    ("Gender Fail Rate Gap", '=MAX(E33:E34)-MIN(E33:E34)', "MAX-MIN", pct_fmt),
    ("Max Fail Rate (Ethnicity)", '=MAX(G42:G45)', "MAX", pct_fmt),
    ("Min Fail Rate (Ethnicity)", '=MIN(G42:G45)', "MIN", pct_fmt),
    ("Ethnicity Fail Rate Gap", '=MAX(G42:G45)-MIN(G42:G45)', "MAX-MIN", pct_fmt),
    ("Max Avg Risk (Education)", '=MAX(F50:F54)', "MAX", dec_fmt),
    ("Min Avg Risk (Education)", '=MIN(F50:F54)', "MIN", dec_fmt),
    ("Education Risk Gap", '=MAX(F50:F54)-MIN(F50:F54)', "MAX-MIN", dec_fmt),
    ("Overall Bias Flag", '=IF(MAX(E33:E34)-MIN(E33:E34)>0.05,"REVIEW NEEDED","ACCEPTABLE")', "IF(gap>5%)", None),
]
for idx, (dlbl, dfml, dexp, dfmt) in enumerate(disparity_metrics):
    r = 58 + idx
    sc(ws5, r, 2, dlbl, font=bold_font, align=Alignment(horizontal="left"))
    sc(ws5, r, 3, dfml, fmt=dfmt, align=Alignment(horizontal="center"))
    sc(ws5, r, 4, dexp, font=Font(name="Calibri", italic=True, size=9, color="666666"), align=Alignment(horizontal="left"))
stripe(ws5, 58, 67, 2, 4)

# Fairness chart
ch_fair = BarChart()
ch_fair.type = "col"; ch_fair.style = 10
ch_fair.title = "Fail Rate by Demographic Group"
ch_fair.y_axis.title = "Fail Rate"; ch_fair.y_axis.numFmt = '0%'
cats_fair = Reference(ws5, min_col=2, min_row=33, max_row=38)
vals_fair = Reference(ws5, min_col=5, min_row=32, max_row=38)
ch_fair.add_data(vals_fair, titles_from_data=True)
ch_fair.set_categories(cats_fair)
ch_fair.width = 18; ch_fair.height = 11
ch_fair.series[0].graphicalProperties.solidFill = MED_BLUE
ch_fair.legend = None
dl_fair = DataLabelList(); dl_fair.showVal = True; dl_fair.numFmt = '0.0%'
ch_fair.series[0].dLbls = dl_fair
ws5.add_chart(ch_fair, "H24")

for c in range(1, 12):
    ws5.column_dimensions[get_column_letter(c)].width = 20

# ============================================================
# FINAL SETTINGS
# ============================================================
tab_colors = ["1B2A4A","2E5090","E8A838","70AD47","C00000"]
sheets = [ws1, ws2, ws3, ws4, ws5]
for ws, color in zip(sheets, tab_colors):
    ws.sheet_properties.tabColor = color
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

wb.active = wb.sheetnames.index("1. Academic Overview")
wb.save(OUT_PATH)
print("=" * 60)
print("  Dashboard saved:", OUT_PATH)
print("  5 sheets with Excel formulas + charts")
print("  " + str(N) + " students processed")
print("=" * 60)
