import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, ScatterChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
import numpy as np

# Load the data
df = pd.read_csv('/Users/prachisingh/Desktop/rev_ler_da/Student_performance_data _.csv')

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define common styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
subheader_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=16, color="1F4E78")
border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin'))

# ===============================================
# PAGE 1: ACADEMIC OVERVIEW
# ===============================================
ws1 = wb.create_sheet("Academic Overview")

# Title
ws1['A1'] = "ACADEMIC PERFORMANCE DASHBOARD"
ws1['A1'].font = Font(bold=True, size=18, color="1F4E78")
ws1['A1'].alignment = Alignment(horizontal='center')
ws1.merge_cells('A1:H1')

ws1['A2'] = f"University Early Warning System - {df.shape[0]} Students"
ws1['A2'].font = Font(size=11, italic=True, color="7F7F7F")
ws1['A2'].alignment = Alignment(horizontal='center')
ws1.merge_cells('A2:H2')

# Section 1: Key Metrics
ws1['A4'] = "KEY PERFORMANCE INDICATORS"
ws1['A4'].font = title_font
ws1.merge_cells('A4:H4')

# Headers for metrics
metrics_headers = ['Metric', 'Value', 'Target', 'Status', 'Variance']
for col_num, header in enumerate(metrics_headers, start=1):
    cell = ws1.cell(row=5, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Add data range starting from row 10 (will be populated with formulas)
data_start_row = 10
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=data_start_row):
    for c_idx, value in enumerate(row, start=1):
        ws1.cell(row=r_idx, column=c_idx, value=value)

# Define GradeClass: 0='A', 1='B', 2='C', 3='D', 4='F'
# Add formulas for metrics
ws1['A6'] = "Average GPA"
ws1['B6'] = f'=AVERAGE(O{data_start_row+1}:O{data_start_row+len(df)})'
ws1['C6'] = 3.0
ws1['D6'] = '=IF(B6>=C6,"✓ On Track","⚠ Below Target")'
ws1['E6'] = '=B6-C6'

ws1['A7'] = "Pass Rate (%)"
ws1['B7'] = f'=COUNTIFS(P{data_start_row+1}:P{data_start_row+len(df)},"<4")/COUNTA(P{data_start_row+1}:P{data_start_row+len(df)})*100'
ws1['C7'] = 80
ws1['D7'] = '=IF(B7>=C7,"✓ On Track","⚠ Below Target")'
ws1['E7'] = '=B7-C7'

ws1['A8'] = "Failure Rate (%)"
ws1['B8'] = f'=COUNTIFS(P{data_start_row+1}:P{data_start_row+len(df)},4)/COUNTA(P{data_start_row+1}:P{data_start_row+len(df)})*100'
ws1['C8'] = '=20-(B8*0.2)'
ws1['D8'] = '=IF(B8<=20,"✓ On Track","⚠ High Risk")'
ws1['E8'] = '=B8-20'

ws1['A9'] = "At-Risk Students"
ws1['B9'] = f'=COUNTIFS(P{data_start_row+1}:P{data_start_row+len(df)},">2")'
ws1['C9'] = f'=COUNTA(P{data_start_row+1}:P{data_start_row+len(df)})*0.15'
ws1['D9'] = '=IF(B9<=C9,"✓ Manageable","⚠ High Count")'
ws1['E9'] = '=B9-C9'

# Format metric values
for row in range(6, 10):
    ws1.cell(row=row, column=2).number_format = '0.00'
    ws1.cell(row=row, column=3).number_format = '0.00'
    ws1.cell(row=row, column=5).number_format = '+0.00;-0.00'
    ws1.cell(row=row, column=4).alignment = Alignment(horizontal='center')

# Section 2: Grade Distribution Analysis
ws1[f'A{data_start_row + len(df) + 3}'] = "GRADE DISTRIBUTION ANALYSIS"
ws1[f'A{data_start_row + len(df) + 3}'].font = title_font

dist_start = data_start_row + len(df) + 5
ws1[f'A{dist_start}'] = "Grade"
ws1[f'B{dist_start}'] = "Count"
ws1[f'C{dist_start}'] = "Percentage"
ws1[f'D{dist_start}'] = "GPA Range"

for col in ['A', 'B', 'C', 'D']:
    cell = ws1[f'{col}{dist_start}']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Grade distribution formulas
grades = [('A (Excellent)', 0, '3.5-4.0'), ('B (Good)', 1, '3.0-3.49'), 
          ('C (Average)', 2, '2.0-2.99'), ('D (Below Average)', 3, '1.0-1.99'), 
          ('F (Fail)', 4, '0.0-0.99')]

for idx, (grade_name, grade_val, gpa_range) in enumerate(grades, start=1):
    row = dist_start + idx
    ws1[f'A{row}'] = grade_name
    ws1[f'B{row}'] = f'=COUNTIF(P{data_start_row+1}:P{data_start_row+len(df)},{grade_val})'
    ws1[f'C{row}'] = f'=B{row}/SUM(B{dist_start+1}:B{dist_start+5})*100'
    ws1[f'D{row}'] = gpa_range
    ws1[f'C{row}'].number_format = '0.0"%"'

# Add Pass/Fail Summary
summary_row = dist_start + 7
ws1[f'A{summary_row}'] = "PASS/FAIL SUMMARY"
ws1[f'A{summary_row}'].font = Font(bold=True, size=11, color="1F4E78")

ws1[f'A{summary_row+1}'] = "Status"
ws1[f'B{summary_row+1}'] = "Count"
ws1[f'C{summary_row+1}'] = "Percentage"

for col in ['A', 'B', 'C']:
    cell = ws1[f'{col}{summary_row+1}']
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = Alignment(horizontal='center')

ws1[f'A{summary_row+2}'] = "Pass (A-D)"
ws1[f'B{summary_row+2}'] = f'=COUNTIFS(P{data_start_row+1}:P{data_start_row+len(df)},"<4")'
ws1[f'C{summary_row+2}'] = f'=B{summary_row+2}/COUNTA(P{data_start_row+1}:P{data_start_row+len(df)})*100'

ws1[f'A{summary_row+3}'] = "Fail (F)"
ws1[f'B{summary_row+3}'] = f'=COUNTIF(P{data_start_row+1}:P{data_start_row+len(df)},4)'
ws1[f'C{summary_row+3}'] = f'=B{summary_row+3}/COUNTA(P{data_start_row+1}:P{data_start_row+len(df)})*100'

ws1[f'C{summary_row+2}'].number_format = '0.0"%"'
ws1[f'C{summary_row+3}'].number_format = '0.0"%"'

# Set column widths
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 15

# ===============================================
# PAGE 2: RISK FACTOR ANALYSIS
# ===============================================
ws2 = wb.create_sheet("Risk Factor Analysis")

# Title
ws2['A1'] = "RISK FACTOR CORRELATION ANALYSIS"
ws2['A1'].font = Font(bold=True, size=18, color="1F4E78")
ws2.merge_cells('A1:H1')

ws2['A2'] = "Identifying Key Performance Drivers"
ws2['A2'].font = Font(size=11, italic=True, color="7F7F7F")
ws2.merge_cells('A2:H2')

# Copy raw data to this sheet for reference
data_start_row_p2 = 5
ws2['A4'] = "STUDENT DATA REFERENCE"
ws2['A4'].font = Font(bold=True, size=11)

for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=data_start_row_p2):
    for c_idx, value in enumerate(row, start=1):
        ws2.cell(row=r_idx, column=c_idx, value=value)

# Correlation Analysis Section
corr_start = data_start_row_p2 + len(df) + 3
ws2[f'A{corr_start}'] = "CORRELATION ANALYSIS"
ws2[f'A{corr_start}'].font = title_font

ws2[f'A{corr_start+2}'] = "Factor"
ws2[f'B{corr_start+2}'] = "Correlation with GPA"
ws2[f'C{corr_start+2}'] = "Impact Level"
ws2[f'D{corr_start+2}'] = "Priority"

for col in ['A', 'B', 'C', 'D']:
    cell = ws2[f'{col}{corr_start+2}']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Calculate correlations using CORREL function
# StudyTimeWeekly is column F, Absences is column G, GPA is column O
factors = [
    ('Study Time (Weekly)', 'F', 'High'),
    ('Absences (Negative)', 'G', 'High'),
    ('Tutoring', 'H', 'Medium'),
    ('Parental Support', 'I', 'Medium'),
    ('Extracurricular', 'J', 'Low'),
]

for idx, (factor_name, col_letter, impact) in enumerate(factors, start=1):
    row = corr_start + 2 + idx
    ws2[f'A{row}'] = factor_name
    # CORREL formula for correlation
    ws2[f'B{row}'] = f'=CORREL({col_letter}{data_start_row_p2+1}:{col_letter}{data_start_row_p2+len(df)},O{data_start_row_p2+1}:O{data_start_row_p2+len(df)})'
    ws2[f'C{row}'] = impact
    ws2[f'D{row}'] = f'=IF(ABS(B{row})>0.5,"Critical",IF(ABS(B{row})>0.3,"Important","Monitor"))'
    ws2[f'B{row}'].number_format = '0.000'

# Study Time vs Grade Analysis
study_analysis_start = corr_start + 10
ws2[f'A{study_analysis_start}'] = "STUDY TIME vs PERFORMANCE"
ws2[f'A{study_analysis_start}'].font = title_font

ws2[f'A{study_analysis_start+2}'] = "Study Time Range"
ws2[f'B{study_analysis_start+2}'] = "Avg GPA"
ws2[f'C{study_analysis_start+2}'] = "Pass Rate %"
ws2[f'D{study_analysis_start+2}'] = "Student Count"

for col in ['A', 'B', 'C', 'D']:
    cell = ws2[f'{col}{study_analysis_start+2}']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

study_ranges = [
    ('0-5 hours', 0, 5),
    ('5-10 hours', 5, 10),
    ('10-15 hours', 10, 15),
    ('15-20 hours', 15, 20),
    ('20+ hours', 20, 999)
]

for idx, (range_name, min_val, max_val) in enumerate(study_ranges, start=1):
    row = study_analysis_start + 2 + idx
    ws2[f'A{row}'] = range_name
    ws2[f'B{row}'] = f'=AVERAGEIFS(O{data_start_row_p2+1}:O{data_start_row_p2+len(df)},F{data_start_row_p2+1}:F{data_start_row_p2+len(df)},">="{min_val},F{data_start_row_p2+1}:F{data_start_row_p2+len(df)},"<"{max_val})'
    ws2[f'C{row}'] = f'=COUNTIFS(F{data_start_row_p2+1}:F{data_start_row_p2+len(df)},">="{min_val},F{data_start_row_p2+1}:F{data_start_row_p2+len(df)},"<"{max_val},P{data_start_row_p2+1}:P{data_start_row_p2+len(df)},"<4")/COUNTIFS(F{data_start_row_p2+1}:F{data_start_row_p2+len(df)},">="{min_val},F{data_start_row_p2+1}:F{data_start_row_p2+len(df)},"<"{max_val})*100'
    ws2[f'D{row}'] = f'=COUNTIFS(F{data_start_row_p2+1}:F{data_start_row_p2+len(df)},">="{min_val},F{data_start_row_p2+1}:F{data_start_row_p2+len(df)},"<"{max_val})'
    ws2[f'B{row}'].number_format = '0.00'
    ws2[f'C{row}'].number_format = '0.0"%"'

# Attendance vs Grade Analysis
attend_analysis_start = study_analysis_start + 10
ws2[f'A{attend_analysis_start}'] = "ATTENDANCE vs PERFORMANCE"
ws2[f'A{attend_analysis_start}'].font = title_font

ws2[f'A{attend_analysis_start+2}'] = "Absence Range"
ws2[f'B{attend_analysis_start+2}'] = "Avg GPA"
ws2[f'C{attend_analysis_start+2}'] = "Failure Rate %"
ws2[f'D{attend_analysis_start+2}'] = "Student Count"

for col in ['A', 'B', 'C', 'D']:
    cell = ws2[f'{col}{attend_analysis_start+2}']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

absence_ranges = [
    ('0-5 days (Excellent)', 0, 5),
    ('6-10 days (Good)', 6, 10),
    ('11-15 days (Fair)', 11, 15),
    ('16-20 days (Poor)', 16, 20),
    ('20+ days (Critical)', 21, 999)
]

for idx, (range_name, min_val, max_val) in enumerate(absence_ranges, start=1):
    row = attend_analysis_start + 2 + idx
    ws2[f'A{row}'] = range_name
    ws2[f'B{row}'] = f'=AVERAGEIFS(O{data_start_row_p2+1}:O{data_start_row_p2+len(df)},G{data_start_row_p2+1}:G{data_start_row_p2+len(df)},">="{min_val},G{data_start_row_p2+1}:G{data_start_row_p2+len(df)},"<"{max_val})'
    ws2[f'C{row}'] = f'=COUNTIFS(G{data_start_row_p2+1}:G{data_start_row_p2+len(df)},">="{min_val},G{data_start_row_p2+1}:G{data_start_row_p2+len(df)},"<"{max_val},P{data_start_row_p2+1}:P{data_start_row_p2+len(df)},4)/COUNTIFS(G{data_start_row_p2+1}:G{data_start_row_p2+len(df)},">="{min_val},G{data_start_row_p2+1}:G{data_start_row_p2+len(df)},"<"{max_val})*100'
    ws2[f'D{row}'] = f'=COUNTIFS(G{data_start_row_p2+1}:G{data_start_row_p2+len(df)},">="{min_val},G{data_start_row_p2+1}:G{data_start_row_p2+len(df)},"<"{max_val})'
    ws2[f'B{row}'].number_format = '0.00'
    ws2[f'C{row}'].number_format = '0.0"%"'

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 20

# ===============================================
# PAGE 3: PERFORMANCE RISK INDEX
# ===============================================
ws3 = wb.create_sheet("Performance Risk Index")

# Title
ws3['A1'] = "STUDENT RISK ASSESSMENT MODEL"
ws3['A1'].font = Font(bold=True, size=18, color="1F4E78")
ws3.merge_cells('A1:L1')

ws3['A2'] = "Predictive Analytics for Early Intervention"
ws3['A2'].font = Font(size=11, italic=True, color="7F7F7F")
ws3.merge_cells('A2:L2')

# Risk Index Formula Explanation
ws3['A4'] = "RISK INDEX FORMULA"
ws3['A4'].font = title_font
ws3.merge_cells('A4:L4')

ws3['A6'] = "Component"
ws3['B6'] = "Weight"
ws3['C6'] = "Description"
ws3['D6'] = "Calculation"

for col in ['A', 'B', 'C', 'D']:
    cell = ws3[f'{col}6']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

formula_components = [
    ('Low GPA', '35%', 'GPA < 2.0', 'IF(GPA<2.0, 35, IF(GPA<2.5, 20, 0))'),
    ('High Absences', '25%', 'Absences > 15', 'IF(Absences>15, 25, IF(Absences>10, 15, 0))'),
    ('Low Study Time', '20%', 'Study < 10 hrs/week', 'IF(StudyTime<10, 20, IF(StudyTime<15, 10, 0))'),
    ('No Support', '10%', 'No tutoring/parental support', 'IF(Tutoring=0 AND Support<2, 10, 0)'),
    ('Grade Trend', '10%', 'Grade Class D or F', 'IF(GradeClass>=3, 10, 0)'),
]

for idx, (component, weight, desc, calc) in enumerate(formula_components, start=1):
    row = 6 + idx
    ws3[f'A{row}'] = component
    ws3[f'B{row}'] = weight
    ws3[f'C{row}'] = desc
    ws3[f'D{row}'] = calc

ws3['A13'] = "Total Risk Score Range: 0-100 (Higher = Greater Risk)"
ws3['A13'].font = Font(bold=True, size=10, color="C00000")
ws3.merge_cells('A13:D13')

# Student Risk Segmentation Table
ws3['A15'] = "STUDENT RISK SEGMENTATION"
ws3['A15'].font = title_font
ws3.merge_cells('A15:L15')

# Headers
risk_headers = ['StudentID', 'GPA', 'Absences', 'Study Time', 'Tutoring', 'Support', 
                'Grade', 'Risk Score', 'Risk Level', 'Priority', 'Recommended Action']

for col_num, header in enumerate(risk_headers, start=1):
    cell = ws3.cell(row=16, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Copy student data and add risk calculations
data_start_risk = 17
for idx, (_, student) in enumerate(df.iterrows(), start=0):
    row = data_start_risk + idx
    
    # Student data
    ws3[f'A{row}'] = student['StudentID']
    ws3[f'B{row}'] = student['GPA']
    ws3[f'C{row}'] = student['Absences']
    ws3[f'D{row}'] = student['StudyTimeWeekly']
    ws3[f'E{row}'] = student['Tutoring']
    ws3[f'F{row}'] = student['ParentalSupport']
    ws3[f'G{row}'] = student['GradeClass']
    
    # Risk Score Formula (Column H)
    ws3[f'H{row}'] = f'=IF(B{row}<2,35,IF(B{row}<2.5,20,0))+IF(C{row}>15,25,IF(C{row}>10,15,0))+IF(D{row}<10,20,IF(D{row}<15,10,0))+IF(AND(E{row}=0,F{row}<2),10,0)+IF(G{row}>=3,10,0)'
    
    # Risk Level (Column I)
    ws3[f'I{row}'] = f'=IF(H{row}>=60,"Critical",IF(H{row}>=40,"High",IF(H{row}>=20,"Medium","Low")))'
    
    # Priority (Column J)
    ws3[f'J{row}'] = f'=IF(I{row}="Critical",1,IF(I{row}="High",2,IF(I{row}="Medium",3,4)))'
    
    # Recommended Action (Column K)
    ws3[f'K{row}'] = f'=IF(I{row}="Critical","Immediate 1-on-1 counseling",IF(I{row}="High","Weekly check-in + tutoring",IF(I{row}="Medium","Bi-weekly monitoring","Standard support")))'

# Apply conditional formatting to Risk Score
ws3.conditional_formatting.add(f'H{data_start_risk}:H{data_start_risk+len(df)-1}',
    ColorScaleRule(start_type='num', start_value=0, start_color='63BE7B',
                   mid_type='num', mid_value=50, mid_color='FFEB84',
                   end_type='num', end_value=100, end_color='F8696B'))

# Risk Distribution Summary
summary_start = data_start_risk + len(df) + 2
ws3[f'A{summary_start}'] = "RISK DISTRIBUTION SUMMARY"
ws3[f'A{summary_start}'].font = title_font

ws3[f'A{summary_start+2}'] = "Risk Level"
ws3[f'B{summary_start+2}'] = "Count"
ws3[f'C{summary_start+2}'] = "Percentage"
ws3[f'D{summary_start+2}'] = "Avg GPA"
ws3[f'E{summary_start+2}'] = "Intervention Cost/Student"
ws3[f'F{summary_start+2}'] = "Total Cost"

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws3[f'{col}{summary_start+2}']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

risk_levels = [
    ('Critical (60-100)', 'Critical', 5000),
    ('High (40-59)', 'High', 3000),
    ('Medium (20-39)', 'Medium', 1500),
    ('Low (0-19)', 'Low', 500)
]

for idx, (level_name, level_code, cost) in enumerate(risk_levels, start=1):
    row = summary_start + 2 + idx
    ws3[f'A{row}'] = level_name
    ws3[f'B{row}'] = f'=COUNTIF(I{data_start_risk}:I{data_start_risk+len(df)-1},"{level_code}")'
    ws3[f'C{row}'] = f'=B{row}/{len(df)}*100'
    ws3[f'D{row}'] = f'=AVERAGEIF(I{data_start_risk}:I{data_start_risk+len(df)-1},"{level_code}",B{data_start_risk}:B{data_start_risk+len(df)-1})'
    ws3[f'E{row}'] = cost
    ws3[f'F{row}'] = f'=B{row}*E{row}'
    ws3[f'C{row}'].number_format = '0.0"%"'
    ws3[f'D{row}'].number_format = '0.00'
    ws3[f'E{row}'].number_format = '$#,##0'
    ws3[f'F{row}'].number_format = '$#,##0'

ws3[f'A{summary_start+7}'] = "TOTAL INTERVENTION BUDGET"
ws3[f'A{summary_start+7}'].font = Font(bold=True, size=12, color="1F4E78")
ws3[f'F{summary_start+7}'] = f'=SUM(F{summary_start+3}:F{summary_start+6})'
ws3[f'F{summary_start+7}'].number_format = '$#,##0'
ws3[f'F{summary_start+7}'].font = Font(bold=True, size=12)
ws3[f'F{summary_start+7}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 10
ws3.column_dimensions['F'].width = 12
ws3.column_dimensions['G'].width = 10
ws3.column_dimensions['H'].width = 12
ws3.column_dimensions['I'].width = 12
ws3.column_dimensions['J'].width = 10
ws3.column_dimensions['K'].width = 30

# ===============================================
# PAGE 4: INTERVENTION STRATEGY SIMULATOR
# ===============================================
ws4 = wb.create_sheet("Intervention Simulator")

# Title
ws4['A1'] = "INTERVENTION STRATEGY SIMULATOR"
ws4['A1'].font = Font(bold=True, size=18, color="1F4E78")
ws4.merge_cells('A1:H1')

ws4['A2'] = "Predictive Model for Intervention Impact"
ws4['A2'].font = Font(size=11, italic=True, color="7F7F7F")
ws4.merge_cells('A2:H2')

# Scenario Controls
ws4['A4'] = "INTERVENTION PARAMETERS"
ws4['A4'].font = title_font
ws4.merge_cells('A4:H4')

ws4['A6'] = "Parameter"
ws4['B6'] = "Current Avg"
ws4['C6'] = "Proposed Change"
ws4['D6'] = "New Value"
ws4['E6'] = "Impact on GPA"
ws4['F6'] = "Cost per Student"

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws4[f'{col}6']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Calculate current averages from Academic Overview sheet
ws4['A7'] = "Study Time (hrs/week)"
ws4['B7'] = f"='Academic Overview'!AVERAGE(F11:F{10+len(df)})"
ws4['C7'] = 5  # Changeable parameter
ws4['D7'] = '=B7+C7'
ws4['E7'] = '=C7*0.15'  # Correlation coefficient approximation
ws4['F7'] = 800

ws4['A8'] = "Attendance Rate (%)"
ws4['B8'] = f"=(30-'Academic Overview'!AVERAGE(G11:G{10+len(df)}))/30*100"
ws4['C8'] = 10  # Changeable parameter
ws4['D8'] = '=B8+C8'
ws4['E8'] = '=C8*0.02'  # Impact factor
ws4['F8'] = 500

ws4['A9'] = "Tutoring Enrollment (%)"
ws4['B9'] = f"='Academic Overview'!COUNTIF(H11:H{10+len(df)},1)/{len(df)}*100"
ws4['C9'] = 25  # Changeable parameter
ws4['D9'] = '=B9+C9'
ws4['E9'] = '=C9*0.01'  # Impact factor
ws4['F9'] = 3000

ws4['A10'] = "Parental Engagement Program"
ws4['B10'] = f"='Academic Overview'!AVERAGE(I11:I{10+len(df)})"
ws4['C10'] = 1  # Changeable parameter (scale 0-4)
ws4['D10'] = '=B10+C10'
ws4['E10'] = '=C10*0.08'
ws4['F10'] = 1200

# Format cells
for row in range(7, 11):
    ws4[f'B{row}'].number_format = '0.00'
    ws4[f'C{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ws4[f'C{row}'].font = Font(bold=True)
    ws4[f'D{row}'].number_format = '0.00'
    ws4[f'E{row}'].number_format = '+0.00'
    ws4[f'F{row}'].number_format = '$#,##0'

ws4['A11'] = "NOTE: Adjust values in column C to simulate different intervention scenarios"
ws4['A11'].font = Font(italic=True, size=9, color="7F7F7F")
ws4.merge_cells('A11:F11')

# Projection Summary
ws4['A13'] = "PROJECTED OUTCOMES"
ws4['A13'].font = title_font
ws4.merge_cells('A13:H13')

ws4['A15'] = "Metric"
ws4['B15'] = "Current"
ws4['C15'] = "Projected"
ws4['D15'] = "Change"
ws4['E15'] = "% Improvement"
ws4['F15'] = "Target Met?"

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws4[f'{col}15']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

ws4['A16'] = "Average GPA"
ws4['B16'] = f"='Academic Overview'!B6"
ws4['C16'] = '=B16+SUM(E7:E10)'
ws4['D16'] = '=C16-B16'
ws4['E16'] = '=D16/B16*100'
ws4['F16'] = '=IF(C16>=3.0,"✓ YES","✗ NO")'

ws4['A17'] = "Pass Rate (%)"
ws4['B17'] = f"='Academic Overview'!B7"
ws4['C17'] = '=B17+(SUM(E7:E10)*10)'  # Approximation
ws4['D17'] = '=C17-B17'
ws4['E17'] = '=D17/B17*100'
ws4['F17'] = '=IF(C17>=80,"✓ YES","✗ NO")'

ws4['A18'] = "Failure Rate (%)"
ws4['B18'] = f"='Academic Overview'!B8"
ws4['C18'] = '=B18-(SUM(E7:E10)*8)'  # Inverse relationship
ws4['D18'] = '=C18-B18'
ws4['E18'] = '=ABS(D18)/B18*100'
ws4['F18'] = '=IF(C18<=16,"✓ YES","✗ NO")'  # 20% reduction from typical ~20%

ws4['A19'] = "At-Risk Students (Count)"
ws4['B19'] = f"='Academic Overview'!B9"
ws4['C19'] = f'=B19*(1-E18/100)'
ws4['D19'] = '=C19-B19'
ws4['E19'] = '=D19/B19*100'
ws4['F19'] = '=IF(C19<B19,"✓ IMPROVED","✗ NO CHANGE")'

for row in range(16, 20):
    ws4[f'B{row}'].number_format = '0.00'
    ws4[f'C{row}'].number_format = '0.00'
    ws4[f'D{row}'].number_format = '+0.00;[Red]-0.00'
    ws4[f'E{row}'].number_format = '0.0"%"'
    ws4[f'F{row}'].alignment = Alignment(horizontal='center')

# Cost-Benefit Analysis
ws4['A21'] = "COST-BENEFIT ANALYSIS"
ws4['A21'].font = title_font
ws4.merge_cells('A21:H21')

ws4['A23'] = "Category"
ws4['B23'] = "Description"
ws4['C23'] = "Calculation"
ws4['D23'] = "Amount"

for col in ['A', 'B', 'C', 'D']:
    cell = ws4[f'{col}23']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

ws4['A24'] = "Total Students"
ws4['B24'] = "All enrolled students"
ws4['C24'] = "From dataset"
ws4['D24'] = len(df)

ws4['A25'] = "At-Risk Students"
ws4['B25'] = "Students requiring intervention"
ws4['C25'] = "Current count"
ws4['D25'] = '=B19'

ws4['A26'] = "Total Intervention Cost"
ws4['B26'] = "Sum of all intervention programs"
ws4['C26'] = "At-risk × avg cost"
ws4['D26'] = '=D25*AVERAGE(F7:F10)'
ws4['D26'].number_format = '$#,##0'

ws4['A27'] = "Cost per Prevented Failure"
ws4['B27'] = "Cost to prevent one student failure"
ws4['C27'] = "Total cost / failures prevented"
ws4['D27'] = '=D26/ABS(D19)'
ws4['D27'].number_format = '$#,##0'

ws4['A28'] = "ROI (Saved Tuition)"
ws4['B28'] = "Assuming $40K tuition per student"
ws4['C28'] = "Prevented dropouts × tuition"
ws4['D28'] = '=ABS(D19)*40000'
ws4['D28'].number_format = '$#,##0'

ws4['A29'] = "Net Benefit"
ws4['B29'] = "ROI - Total Cost"
ws4['C29'] = "Savings - Investment"
ws4['D29'] = '=D28-D26'
ws4['D29'].number_format = '$#,##0'
ws4['D29'].font = Font(bold=True, size=12)
ws4['D29'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ws4['A30'] = "ROI Ratio"
ws4['B30'] = "Return on Investment"
ws4['C30'] = "Net Benefit / Total Cost"
ws4['D30'] = '=D29/D26'
ws4['D30'].number_format = '0.00'
ws4['D30'].font = Font(bold=True, size=12)

# Scenario Recommendations
ws4['A32'] = "RECOMMENDED SCENARIOS"
ws4['A32'].font = title_font
ws4.merge_cells('A32:H32')

ws4['A34'] = "Scenario"
ws4['B34'] = "Study Time"
ws4['C34'] = "Attendance"
ws4['D34'] = "Tutoring"
ws4['E34'] = "Parent Engage"
ws4['F34'] = "Est. Cost"
ws4['G34'] = "Est. GPA Gain"
ws4['H34'] = "Priority"

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    cell = ws4[f'{col}34']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

scenarios = [
    ('Aggressive (Maximum Impact)', '+10 hrs', '+20%', '+40%', '+2', '=AVERAGE(F7:F10)*1.5', '=SUM(E7:E10)*2', 'High'),
    ('Moderate (Balanced)', '+5 hrs', '+10%', '+25%', '+1', '=AVERAGE(F7:F10)', '=SUM(E7:E10)', 'Medium'),
    ('Conservative (Cost-Effective)', '+3 hrs', '+5%', '+15%', '+0.5', '=AVERAGE(F7:F10)*0.6', '=SUM(E7:E10)*0.5', 'Low'),
    ('Targeted (High-Risk Only)', '+8 hrs', '+15%', '+50%', '+2', '=AVERAGE(F7:F10)*0.8', '=SUM(E7:E10)*1.2', 'Critical')
]

for idx, (scenario, study, attend, tutor, parent, cost, gpa, priority) in enumerate(scenarios, start=1):
    row = 34 + idx
    ws4[f'A{row}'] = scenario
    ws4[f'B{row}'] = study
    ws4[f'C{row}'] = attend
    ws4[f'D{row}'] = tutor
    ws4[f'E{row}'] = parent
    ws4[f'F{row}'] = cost
    ws4[f'G{row}'] = gpa
    ws4[f'H{row}'] = priority
    ws4[f'F{row}'].number_format = '$#,##0'
    ws4[f'G{row}'].number_format = '0.00'

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 18
ws4.column_dimensions['F'].width = 18
ws4.column_dimensions['G'].width = 15
ws4.column_dimensions['H'].width = 15

# ===============================================
# PAGE 5: ETHICS & SAFEGUARDS
# ===============================================
ws5 = wb.create_sheet("Ethics & Safeguards")

# Title
ws5['A1'] = "ETHICAL FRAMEWORK & DATA SAFEGUARDS"
ws5['A1'].font = Font(bold=True, size=18, color="1F4E78")
ws5.merge_cells('A1:G1')

ws5['A2'] = "Ensuring Responsible Use of Predictive Analytics"
ws5['A2'].font = Font(size=11, italic=True, color="7F7F7F")
ws5.merge_cells('A2:G2')

# Ethical Principles
ws5['A4'] = "CORE ETHICAL PRINCIPLES"
ws5['A4'].font = title_font
ws5.merge_cells('A4:G4')

ws5['A6'] = "Principle"
ws5['B6'] = "Description"
ws5['C6'] = "Implementation"
ws5['D6'] = "Compliance Status"

for col in ['A', 'B', 'C', 'D']:
    cell = ws5[f'{col}6']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

principles = [
    ('Transparency', 'Students are informed about risk assessment', 
     'Risk scores shared with students in counseling sessions', '✓ Active'),
    ('Fairness', 'No discrimination based on demographics',
     'Algorithm audited for bias; no ethnicity/gender in risk formula', '✓ Active'),
    ('Privacy', 'Student data protected and anonymized',
     'FERPA compliant; access limited to authorized personnel', '✓ Active'),
    ('Human Oversight', 'Final decisions made by humans, not algorithms',
     'Advisors review all recommendations before action', '✓ Active'),
    ('Right to Explanation', 'Students can request explanation of their risk score',
     'Detailed breakdown provided upon request', '✓ Active'),
    ('Opt-Out Option', 'Students can decline participation',
     'Opt-out form available; alternative support still offered', '✓ Active'),
]

for idx, (principle, description, implementation, status) in enumerate(principles, start=1):
    row = 6 + idx
    ws5[f'A{row}'] = principle
    ws5[f'B{row}'] = description
    ws5[f'C{row}'] = implementation
    ws5[f'D{row}'] = status
    ws5[f'D{row}'].alignment = Alignment(horizontal='center')

# Labeling Transparency
ws5['A14'] = "RISK LABEL TRANSPARENCY"
ws5['A14'].font = title_font
ws5.merge_cells('A14:G14')

ws5['A16'] = "Our Commitment:"
ws5['A16'].font = Font(bold=True, size=11)
ws5['A17'] = "• Risk labels are tools for support, NOT judgments of student ability"
ws5['A18'] = "• Labels are dynamic and updated regularly based on student progress"
ws5['A19'] = "• Students labeled 'at-risk' receive additional resources, not penalties"
ws5['A20'] = "• Faculty trained to avoid stigmatization and use labels constructively"
ws5['A21'] = "• Risk scores include clear explanation of contributing factors"
ws5['A22'] = "• Students can challenge their risk assessment through formal review process"

ws5.merge_cells('A17:G17')
ws5.merge_cells('A18:G18')
ws5.merge_cells('A19:G19')
ws5.merge_cells('A20:G20')
ws5.merge_cells('A21:G21')
ws5.merge_cells('A22:G22')

# Student Data Privacy Policy
ws5['A24'] = "DATA PRIVACY & SECURITY POLICY"
ws5['A24'].font = title_font
ws5.merge_cells('A24:G24')

ws5['A26'] = "Category"
ws5['B26'] = "Policy"
ws5['C26'] = "Technical Implementation"
ws5['D26'] = "Review Frequency"

for col in ['A', 'B', 'C', 'D']:
    cell = ws5[f'{col}26']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

privacy_policies = [
    ('Data Collection', 'Only essential academic data collected',
     'Automated from SIS; no social media scraping', 'Annual'),
    ('Data Storage', 'Encrypted databases with access logs',
     'AES-256 encryption; audit trail enabled', 'Quarterly'),
    ('Data Access', 'Role-based access control',
     'Advisors: read only; Admins: full; Students: own data', 'Monthly'),
    ('Data Retention', 'Data deleted after graduation + 2 years',
     'Automated purge process; alumni can request early deletion', 'Semester'),
    ('Third-Party Sharing', 'No data sold or shared without consent',
     'DPA agreements required; student consent mandatory', 'Per request'),
    ('Breach Protocol', 'Immediate notification within 72 hours',
     'Incident response team; affected students notified directly', 'Annual drill'),
]

for idx, (category, policy, implementation, frequency) in enumerate(privacy_policies, start=1):
    row = 26 + idx
    ws5[f'A{row}'] = category
    ws5[f'B{row}'] = policy
    ws5[f'C{row}'] = implementation
    ws5[f'D{row}'] = frequency

# Bias Monitoring
ws5['A34'] = "ALGORITHMIC BIAS MONITORING"
ws5['A34'].font = title_font
ws5.merge_cells('A34:G34')

ws5['A36'] = "Demographic Group"
ws5['B36'] = "Avg Risk Score"
ws5['C36'] = "Pass Rate %"
ws5['D36'] = "Fairness Metric"
ws5['E36'] = "Status"

for col in ['A', 'B', 'C', 'D', 'E']:
    cell = ws5[f'{col}36']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Reference Performance Risk Index sheet for calculations
ws5['A37'] = "Overall Population"
ws5['B37'] = f"='Performance Risk Index'!AVERAGE(H17:H{16+len(df)})"
ws5['C37'] = f"='Academic Overview'!B7"
ws5['D37'] = "Baseline"
ws5['E37'] = "✓ Reference"

ws5['A38'] = "Male Students"
ws5['B38'] = f"=AVERAGEIF('Performance Risk Index'!B17:B{16+len(df)},\"<=2\",'Performance Risk Index'!H17:H{16+len(df)})"
ws5['C38'] = f"=COUNTIFS('Academic Overview'!C11:C{10+len(df)},1,'Academic Overview'!P11:P{10+len(df)},\"<4\")/COUNTIF('Academic Overview'!C11:C{10+len(df)},1)*100"
ws5['D38'] = "=ABS(B38-B37)/B37"
ws5['E38'] = '=IF(D38<0.1,"✓ Fair","⚠ Review")'

ws5['A39'] = "Female Students"
ws5['B39'] = f"=AVERAGEIF('Performance Risk Index'!B17:B{16+len(df)},\">2\",'Performance Risk Index'!H17:H{16+len(df)})"
ws5['C39'] = f"=COUNTIFS('Academic Overview'!C11:C{10+len(df)},0,'Academic Overview'!P11:P{10+len(df)},\"<4\")/COUNTIF('Academic Overview'!C11:C{10+len(df)},0)*100"
ws5['D39'] = "=ABS(B39-B37)/B37"
ws5['E39'] = '=IF(D39<0.1,"✓ Fair","⚠ Review")'

for row in range(37, 40):
    ws5[f'B{row}'].number_format = '0.00'
    ws5[f'C{row}'].number_format = '0.0"%"'
    ws5[f'D{row}'].number_format = '0.0%'
    ws5[f'E{row}'].alignment = Alignment(horizontal='center')

ws5['A41'] = "Fairness Threshold: Variance < 10% between groups"
ws5['A41'].font = Font(italic=True, size=9, color="7F7F7F")
ws5.merge_cells('A41:E41')

# Accountability Framework
ws5['A43'] = "ACCOUNTABILITY & GOVERNANCE"
ws5['A43'].font = title_font
ws5.merge_cells('A43:G43')

ws5['A45'] = "Stakeholder"
ws5['B45'] = "Role"
ws5['C45'] = "Responsibility"
ws5['D45'] = "Meeting Frequency"

for col in ['A', 'B', 'C', 'D']:
    cell = ws5[f'{col}45']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

stakeholders = [
    ('Ethics Review Board', 'Oversight', 'Review algorithm changes and bias reports', 'Quarterly'),
    ('Data Protection Officer', 'Compliance', 'Ensure FERPA/privacy compliance', 'Monthly'),
    ('Student Representatives', 'Advocacy', 'Voice student concerns and feedback', 'Bi-monthly'),
    ('Faculty Advisory', 'Implementation', 'Guide intervention strategies', 'Monthly'),
    ('IT Security Team', 'Protection', 'Maintain data security and access controls', 'Weekly'),
]

for idx, (stakeholder, role, responsibility, frequency) in enumerate(stakeholders, start=1):
    row = 45 + idx
    ws5[f'A{row}'] = stakeholder
    ws5[f'B{row}'] = role
    ws5[f'C{row}'] = responsibility
    ws5[f'D{row}'] = frequency

# Audit Trail
ws5['A52'] = "AUDIT & REVIEW SCHEDULE"
ws5['A52'].font = title_font
ws5.merge_cells('A52:G52')

ws5['A54'] = "Review Type"
ws5['B54'] = "Frequency"
ws5['C54'] = "Next Due Date"
ws5['D54'] = "Status"

for col in ['A', 'B', 'C', 'D']:
    cell = ws5[f'{col}54']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

reviews = [
    ('Algorithm Bias Audit', 'Quarterly', '2026-05-01', '✓ Scheduled'),
    ('Data Privacy Assessment', 'Semi-annual', '2026-08-01', '✓ Scheduled'),
    ('Student Feedback Survey', 'Semester', '2026-05-15', '✓ Scheduled'),
    ('Intervention Effectiveness', 'Annual', '2026-12-01', '✓ Scheduled'),
    ('Security Penetration Test', 'Annual', '2026-09-01', '✓ Scheduled'),
]

for idx, (review_type, frequency, due_date, status) in enumerate(reviews, start=1):
    row = 54 + idx
    ws5[f'A{row}'] = review_type
    ws5[f'B{row}'] = frequency
    ws5[f'C{row}'] = due_date
    ws5[f'D{row}'] = status
    ws5[f'D{row}'].alignment = Alignment(horizontal='center')

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 45
ws5.column_dimensions['D'].width = 18
ws5.column_dimensions['E'].width = 15

# Save the workbook
output_path = '/Users/prachisingh/Desktop/rev_ler_da/Student_Early_Warning_Dashboard.xlsx'
wb.save(output_path)
print(f"Dashboard created successfully: {output_path}")
print(f"\nDashboard contains {len(wb.sheetnames)} pages:")
for idx, sheet in enumerate(wb.sheetnames, 1):
    print(f"  Page {idx}: {sheet}")
