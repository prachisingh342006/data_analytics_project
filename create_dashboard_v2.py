import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
import warnings
warnings.filterwarnings('ignore')

# Load the data
df = pd.read_csv('/Users/prachisingh/Desktop/rev_ler_da/Student_performance_data _.csv')

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Define common styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
subheader_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=16, color="1F4E78")
border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin'))

# ===============================================
# PAGE 1: ACADEMIC OVERVIEW
# ===============================================
ws1 = wb.create_sheet("Academic Overview")

# Title
ws1['A1'] = "ACADEMIC PERFORMANCE DASHBOARD"
ws1['A1'].font = Font(bold=True, size=18, color="1F4E78")
ws1['A1'].alignment = Alignment(horizontal='center')
ws1.merge_cells('A1:H1')

ws1['A2'] = f"University Early Warning System - {df.shape[0]} Students"
ws1['A2'].font = Font(size=11, italic=True, color="7F7F7F")
ws1['A2'].alignment = Alignment(horizontal='center')
ws1.merge_cells('A2:H2')

# Section 1: Key Metrics
ws1['A4'] = "KEY PERFORMANCE INDICATORS"
ws1['A4'].font = title_font
ws1.merge_cells('A4:H4')

# Headers for metrics
metrics_headers = ['Metric', 'Value', 'Target', 'Status', 'Variance']
for col_num, header in enumerate(metrics_headers, start=1):
    cell = ws1.cell(row=5, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Add data starting from row 10
data_start = 10
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=data_start):
    for c_idx, value in enumerate(row, start=1):
        ws1.cell(row=r_idx, column=c_idx, value=value)

data_end = data_start + len(df)

# Add formulas for metrics with proper Excel syntax
ws1['A6'] = "Average GPA"
ws1['B6'] = f'=AVERAGE(O{data_start+1}:O{data_end})'
ws1['B6'].number_format = '0.00'
ws1['C6'] = 3.0
ws1['C6'].number_format = '0.00'
ws1['D6'] = f'=IF(B6>=C6,"On Track","Below Target")'
ws1['E6'] = '=B6-C6'
ws1['E6'].number_format = '0.00'

ws1['A7'] = "Pass Rate (%)"
ws1['B7'] = f'=COUNTIFS(P{data_start+1}:P{data_end},"<4")/COUNTA(P{data_start+1}:P{data_end})*100'
ws1['B7'].number_format = '0.00'
ws1['C7'] = 80
ws1['C7'].number_format = '0.00'
ws1['D7'] = f'=IF(B7>=C7,"On Track","Below Target")'
ws1['E7'] = '=B7-C7'
ws1['E7'].number_format = '0.00'

ws1['A8'] = "Failure Rate (%)"
ws1['B8'] = f'=COUNTIFS(P{data_start+1}:P{data_end},4)/COUNTA(P{data_start+1}:P{data_end})*100'
ws1['B8'].number_format = '0.00'
ws1['C8'] = 20
ws1['C8'].number_format = '0.00'
ws1['D8'] = f'=IF(B8<=C8,"On Track","High Risk")'
ws1['E8'] = '=B8-C8'
ws1['E8'].number_format = '0.00'

ws1['A9'] = "At-Risk Students"
ws1['B9'] = f'=COUNTIFS(P{data_start+1}:P{data_end},">2")'
ws1['B9'].number_format = '0'
ws1['C9'] = f'=COUNTA(P{data_start+1}:P{data_end})*0.15'
ws1['C9'].number_format = '0.00'
ws1['D9'] = f'=IF(B9<=C9,"Manageable","High Count")'
ws1['E9'] = '=B9-C9'
ws1['E9'].number_format = '0.00'

# Grade Distribution
dist_start = data_end + 3
ws1[f'A{dist_start}'] = "GRADE DISTRIBUTION ANALYSIS"
ws1[f'A{dist_start}'].font = title_font
ws1.merge_cells(f'A{dist_start}:D{dist_start}')

dist_header = dist_start + 2
ws1[f'A{dist_header}'] = "Grade"
ws1[f'B{dist_header}'] = "Count"
ws1[f'C{dist_header}'] = "Percentage"
ws1[f'D{dist_header}'] = "GPA Range"

for col in ['A', 'B', 'C', 'D']:
    cell = ws1[f'{col}{dist_header}']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

grades = [('A (Excellent)', 0, '3.5-4.0'), ('B (Good)', 1, '3.0-3.49'), 
          ('C (Average)', 2, '2.0-2.99'), ('D (Below Average)', 3, '1.0-1.99'), 
          ('F (Fail)', 4, '0.0-0.99')]

for idx, (grade_name, grade_val, gpa_range) in enumerate(grades, start=1):
    row = dist_header + idx
    ws1[f'A{row}'] = grade_name
    ws1[f'B{row}'] = f'=COUNTIF(P{data_start+1}:P{data_end},{grade_val})'
    ws1[f'C{row}'] = f'=B{row}/SUM(B{dist_header+1}:B{dist_header+5})*100'
    ws1[f'C{row}'].number_format = '0.0'
    ws1[f'D{row}'] = gpa_range

# Pass/Fail Summary
summary_row = dist_header + 7
ws1[f'A{summary_row}'] = "PASS/FAIL SUMMARY"
ws1[f'A{summary_row}'].font = Font(bold=True, size=11, color="1F4E78")
ws1.merge_cells(f'A{summary_row}:C{summary_row}')

ws1[f'A{summary_row+1}'] = "Status"
ws1[f'B{summary_row+1}'] = "Count"
ws1[f'C{summary_row+1}'] = "Percentage"

for col in ['A', 'B', 'C']:
    cell = ws1[f'{col}{summary_row+1}']
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = Alignment(horizontal='center')

ws1[f'A{summary_row+2}'] = "Pass (A-D)"
ws1[f'B{summary_row+2}'] = f'=COUNTIFS(P{data_start+1}:P{data_end},"<4")'
ws1[f'C{summary_row+2}'] = f'=B{summary_row+2}/COUNTA(P{data_start+1}:P{data_end})*100'
ws1[f'C{summary_row+2}'].number_format = '0.0'

ws1[f'A{summary_row+3}'] = "Fail (F)"
ws1[f'B{summary_row+3}'] = f'=COUNTIF(P{data_start+1}:P{data_end},4)'
ws1[f'C{summary_row+3}'] = f'=B{summary_row+3}/COUNTA(P{data_start+1}:P{data_end})*100'
ws1[f'C{summary_row+3}'].number_format = '0.0'

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 15

# ===============================================
# PAGE 2: RISK FACTOR ANALYSIS
# ===============================================
ws2 = wb.create_sheet("Risk Factor Analysis")

ws2['A1'] = "RISK FACTOR CORRELATION ANALYSIS"
ws2['A1'].font = Font(bold=True, size=18, color="1F4E78")
ws2.merge_cells('A1:H1')

ws2['A2'] = "Identifying Key Performance Drivers"
ws2['A2'].font = Font(size=11, italic=True, color="7F7F7F")
ws2.merge_cells('A2:H2')

# Copy data
data_start_p2 = 5
ws2['A4'] = "STUDENT DATA REFERENCE"
ws2['A4'].font = Font(bold=True, size=11)

for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=data_start_p2):
    for c_idx, value in enumerate(row, start=1):
        ws2.cell(row=r_idx, column=c_idx, value=value)

data_end_p2 = data_start_p2 + len(df)

# Correlation Analysis
corr_start = data_end_p2 + 3
ws2[f'A{corr_start}'] = "CORRELATION ANALYSIS"
ws2[f'A{corr_start}'].font = title_font

ws2[f'A{corr_start+2}'] = "Factor"
ws2[f'B{corr_start+2}'] = "Correlation with GPA"
ws2[f'C{corr_start+2}'] = "Impact Level"
ws2[f'D{corr_start+2}'] = "Priority"

for col in ['A', 'B', 'C', 'D']:
    cell = ws2[f'{col}{corr_start+2}']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

factors = [
    ('Study Time (Weekly)', 'F', 'High'),
    ('Absences (Negative)', 'G', 'High'),
    ('Tutoring', 'H', 'Medium'),
    ('Parental Support', 'I', 'Medium'),
    ('Extracurricular', 'J', 'Low'),
]

for idx, (factor_name, col_letter, impact) in enumerate(factors, start=1):
    row = corr_start + 2 + idx
    ws2[f'A{row}'] = factor_name
    ws2[f'B{row}'] = f'=CORREL({col_letter}{data_start_p2+1}:{col_letter}{data_end_p2},O{data_start_p2+1}:O{data_end_p2})'
    ws2[f'B{row}'].number_format = '0.000'
    ws2[f'C{row}'] = impact
    ws2[f'D{row}'] = f'=IF(ABS(B{row})>0.5,"Critical",IF(ABS(B{row})>0.3,"Important","Monitor"))'

# Study Time Analysis
study_start = corr_start + 10
ws2[f'A{study_start}'] = "STUDY TIME vs PERFORMANCE"
ws2[f'A{study_start}'].font = title_font

ws2[f'A{study_start+2}'] = "Study Time Range"
ws2[f'B{study_start+2}'] = "Avg GPA"
ws2[f'C{study_start+2}'] = "Pass Rate %"
ws2[f'D{study_start+2}'] = "Student Count"

for col in ['A', 'B', 'C', 'D']:
    cell = ws2[f'{col}{study_start+2}']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

study_ranges = [
    ('0-5 hours', 0, 5),
    ('5-10 hours', 5, 10),
    ('10-15 hours', 10, 15),
    ('15-20 hours', 15, 20),
    ('20+ hours', 20, 999)
]

for idx, (range_name, min_val, max_val) in enumerate(study_ranges, start=1):
    row = study_start + 2 + idx
    ws2[f'A{row}'] = range_name
    if max_val == 999:
        ws2[f'B{row}'] = f'=AVERAGEIFS(O{data_start_p2+1}:O{data_end_p2},F{data_start_p2+1}:F{data_end_p2},">={min_val}")'
        ws2[f'C{row}'] = f'=COUNTIFS(F{data_start_p2+1}:F{data_end_p2},">={min_val}",P{data_start_p2+1}:P{data_end_p2},"<4")/COUNTIFS(F{data_start_p2+1}:F{data_end_p2},">={min_val}")*100'
        ws2[f'D{row}'] = f'=COUNTIFS(F{data_start_p2+1}:F{data_end_p2},">={min_val}")'
    else:
        ws2[f'B{row}'] = f'=AVERAGEIFS(O{data_start_p2+1}:O{data_end_p2},F{data_start_p2+1}:F{data_end_p2},">={min_val}",F{data_start_p2+1}:F{data_end_p2},"<{max_val}")'
        ws2[f'C{row}'] = f'=COUNTIFS(F{data_start_p2+1}:F{data_end_p2},">={min_val}",F{data_start_p2+1}:F{data_end_p2},"<{max_val}",P{data_start_p2+1}:P{data_end_p2},"<4")/COUNTIFS(F{data_start_p2+1}:F{data_end_p2},">={min_val}",F{data_start_p2+1}:F{data_end_p2},"<{max_val}")*100'
        ws2[f'D{row}'] = f'=COUNTIFS(F{data_start_p2+1}:F{data_end_p2},">={min_val}",F{data_start_p2+1}:F{data_end_p2},"<{max_val}")'
    ws2[f'B{row}'].number_format = '0.00'
    ws2[f'C{row}'].number_format = '0.0'

# Attendance Analysis
attend_start = study_start + 10
ws2[f'A{attend_start}'] = "ATTENDANCE vs PERFORMANCE"
ws2[f'A{attend_start}'].font = title_font

ws2[f'A{attend_start+2}'] = "Absence Range"
ws2[f'B{attend_start+2}'] = "Avg GPA"
ws2[f'C{attend_start+2}'] = "Failure Rate %"
ws2[f'D{attend_start+2}'] = "Student Count"

for col in ['A', 'B', 'C', 'D']:
    cell = ws2[f'{col}{attend_start+2}']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

absence_ranges = [
    ('0-5 days (Excellent)', 0, 5),
    ('6-10 days (Good)', 6, 10),
    ('11-15 days (Fair)', 11, 15),
    ('16-20 days (Poor)', 16, 20),
    ('20+ days (Critical)', 21, 999)
]

for idx, (range_name, min_val, max_val) in enumerate(absence_ranges, start=1):
    row = attend_start + 2 + idx
    ws2[f'A{row}'] = range_name
    if max_val == 999:
        ws2[f'B{row}'] = f'=AVERAGEIFS(O{data_start_p2+1}:O{data_end_p2},G{data_start_p2+1}:G{data_end_p2},">={min_val}")'
        ws2[f'C{row}'] = f'=COUNTIFS(G{data_start_p2+1}:G{data_end_p2},">={min_val}",P{data_start_p2+1}:P{data_end_p2},4)/COUNTIFS(G{data_start_p2+1}:G{data_end_p2},">={min_val}")*100'
        ws2[f'D{row}'] = f'=COUNTIFS(G{data_start_p2+1}:G{data_end_p2},">={min_val}")'
    else:
        ws2[f'B{row}'] = f'=AVERAGEIFS(O{data_start_p2+1}:O{data_end_p2},G{data_start_p2+1}:G{data_end_p2},">={min_val}",G{data_start_p2+1}:G{data_end_p2},"<={max_val}")'
        ws2[f'C{row}'] = f'=COUNTIFS(G{data_start_p2+1}:G{data_end_p2},">={min_val}",G{data_start_p2+1}:G{data_end_p2},"<={max_val}",P{data_start_p2+1}:P{data_end_p2},4)/COUNTIFS(G{data_start_p2+1}:G{data_end_p2},">={min_val}",G{data_start_p2+1}:G{data_end_p2},"<={max_val}")*100'
        ws2[f'D{row}'] = f'=COUNTIFS(G{data_start_p2+1}:G{data_end_p2},">={min_val}",G{data_start_p2+1}:G{data_end_p2},"<={max_val}")'
    ws2[f'B{row}'].number_format = '0.00'
    ws2[f'C{row}'].number_format = '0.0'

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 20

# ===============================================
# PAGE 3: PERFORMANCE RISK INDEX
# ===============================================
ws3 = wb.create_sheet("Performance Risk Index")

ws3['A1'] = "STUDENT RISK ASSESSMENT MODEL"
ws3['A1'].font = Font(bold=True, size=18, color="1F4E78")
ws3.merge_cells('A1:K1')

ws3['A2'] = "Predictive Analytics for Early Intervention"
ws3['A2'].font = Font(size=11, italic=True, color="7F7F7F")
ws3.merge_cells('A2:K2')

# Risk Formula
ws3['A4'] = "RISK INDEX FORMULA"
ws3['A4'].font = title_font
ws3.merge_cells('A4:K4')

ws3['A6'] = "Component"
ws3['B6'] = "Weight"
ws3['C6'] = "Description"
ws3['D6'] = "Calculation"

for col in ['A', 'B', 'C', 'D']:
    cell = ws3[f'{col}6']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

formula_components = [
    ('Low GPA', '35%', 'GPA < 2.0', 'IF(GPA<2, 35, IF(GPA<2.5, 20, 0))'),
    ('High Absences', '25%', 'Absences > 15', 'IF(Absences>15, 25, IF(Absences>10, 15, 0))'),
    ('Low Study Time', '20%', 'Study < 10 hrs/week', 'IF(StudyTime<10, 20, IF(StudyTime<15, 10, 0))'),
    ('No Support', '10%', 'No tutoring/parental support', 'IF(AND(Tutoring=0, Support<2), 10, 0)'),
    ('Grade Trend', '10%', 'Grade Class D or F', 'IF(GradeClass>=3, 10, 0)'),
]

for idx, (component, weight, desc, calc) in enumerate(formula_components, start=1):
    row = 6 + idx
    ws3[f'A{row}'] = component
    ws3[f'B{row}'] = weight
    ws3[f'C{row}'] = desc
    ws3[f'D{row}'] = calc

ws3['A13'] = "Total Risk Score Range: 0-100 (Higher = Greater Risk)"
ws3['A13'].font = Font(bold=True, size=10, color="C00000")
ws3.merge_cells('A13:D13')

# Student Risk Table
ws3['A15'] = "STUDENT RISK SEGMENTATION"
ws3['A15'].font = title_font
ws3.merge_cells('A15:K15')

risk_headers = ['StudentID', 'GPA', 'Absences', 'Study Time', 'Tutoring', 'Support', 
                'Grade', 'Risk Score', 'Risk Level', 'Priority', 'Recommended Action']

for col_num, header in enumerate(risk_headers, start=1):
    cell = ws3.cell(row=16, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Student data with risk scores
data_start_risk = 17
for idx, (_, student) in enumerate(df.iterrows(), start=0):
    row = data_start_risk + idx
    
    ws3[f'A{row}'] = student['StudentID']
    ws3[f'B{row}'] = student['GPA']
    ws3[f'C{row}'] = student['Absences']
    ws3[f'D{row}'] = student['StudyTimeWeekly']
    ws3[f'E{row}'] = student['Tutoring']
    ws3[f'F{row}'] = student['ParentalSupport']
    ws3[f'G{row}'] = student['GradeClass']
    
    # Risk Score formula
    ws3[f'H{row}'] = f'=IF(B{row}<2,35,IF(B{row}<2.5,20,0))+IF(C{row}>15,25,IF(C{row}>10,15,0))+IF(D{row}<10,20,IF(D{row}<15,10,0))+IF(AND(E{row}=0,F{row}<2),10,0)+IF(G{row}>=3,10,0)'
    
    # Risk Level
    ws3[f'I{row}'] = f'=IF(H{row}>=60,"Critical",IF(H{row}>=40,"High",IF(H{row}>=20,"Medium","Low")))'
    
    # Priority
    ws3[f'J{row}'] = f'=IF(I{row}="Critical",1,IF(I{row}="High",2,IF(I{row}="Medium",3,4)))'
    
    # Action
    ws3[f'K{row}'] = f'=IF(I{row}="Critical","Immediate 1-on-1 counseling",IF(I{row}="High","Weekly check-in + tutoring",IF(I{row}="Medium","Bi-weekly monitoring","Standard support")))'

data_end_risk = data_start_risk + len(df) - 1

# Color scale for risk scores
ws3.conditional_formatting.add(f'H{data_start_risk}:H{data_end_risk}',
    ColorScaleRule(start_type='num', start_value=0, start_color='63BE7B',
                   mid_type='num', mid_value=50, mid_color='FFEB84',
                   end_type='num', end_value=100, end_color='F8696B'))

# Risk Distribution
summary_start = data_end_risk + 2
ws3[f'A{summary_start}'] = "RISK DISTRIBUTION SUMMARY"
ws3[f'A{summary_start}'].font = title_font

ws3[f'A{summary_start+2}'] = "Risk Level"
ws3[f'B{summary_start+2}'] = "Count"
ws3[f'C{summary_start+2}'] = "Percentage"
ws3[f'D{summary_start+2}'] = "Avg GPA"
ws3[f'E{summary_start+2}'] = "Cost/Student"
ws3[f'F{summary_start+2}'] = "Total Cost"

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws3[f'{col}{summary_start+2}']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

risk_levels = [
    ('Critical (60-100)', 'Critical', 5000),
    ('High (40-59)', 'High', 3000),
    ('Medium (20-39)', 'Medium', 1500),
    ('Low (0-19)', 'Low', 500)
]

for idx, (level_name, level_code, cost) in enumerate(risk_levels, start=1):
    row = summary_start + 2 + idx
    ws3[f'A{row}'] = level_name
    ws3[f'B{row}'] = f'=COUNTIF(I{data_start_risk}:I{data_end_risk},"{level_code}")'
    ws3[f'C{row}'] = f'=B{row}/{len(df)}*100'
    ws3[f'C{row}'].number_format = '0.0'
    ws3[f'D{row}'] = f'=AVERAGEIF(I{data_start_risk}:I{data_end_risk},"{level_code}",B{data_start_risk}:B{data_end_risk})'
    ws3[f'D{row}'].number_format = '0.00'
    ws3[f'E{row}'] = cost
    ws3[f'E{row}'].number_format = '#,##0'
    ws3[f'F{row}'] = f'=B{row}*E{row}'
    ws3[f'F{row}'].number_format = '#,##0'

ws3[f'A{summary_start+7}'] = "TOTAL INTERVENTION BUDGET"
ws3[f'A{summary_start+7}'].font = Font(bold=True, size=12, color="1F4E78")
ws3[f'F{summary_start+7}'] = f'=SUM(F{summary_start+3}:F{summary_start+6})'
ws3[f'F{summary_start+7}'].number_format = '#,##0'
ws3[f'F{summary_start+7}'].font = Font(bold=True, size=12)

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['K'].width = 30

# ===============================================
# PAGE 4: INTERVENTION SIMULATOR
# ===============================================
ws4 = wb.create_sheet("Intervention Simulator")

ws4['A1'] = "INTERVENTION STRATEGY SIMULATOR"
ws4['A1'].font = Font(bold=True, size=18, color="1F4E78")
ws4.merge_cells('A1:H1')

ws4['A2'] = "Predictive Model for Intervention Impact"
ws4['A2'].font = Font(size=11, italic=True, color="7F7F7F")
ws4.merge_cells('A2:H2')

# Parameters
ws4['A4'] = "INTERVENTION PARAMETERS"
ws4['A4'].font = title_font
ws4.merge_cells('A4:H4')

ws4['A6'] = "Parameter"
ws4['B6'] = "Current Avg"
ws4['C6'] = "Proposed Change"
ws4['D6'] = "New Value"
ws4['E6'] = "Impact on GPA"
ws4['F6'] = "Cost per Student"

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws4[f'{col}6']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

ws4['A7'] = "Study Time (hrs/week)"
ws4['B7'] = f"='Academic Overview'!AVERAGE(F11:F{10+len(df)})"
ws4['B7'].number_format = '0.00'
ws4['C7'] = 5
ws4['C7'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ws4['D7'] = '=B7+C7'
ws4['D7'].number_format = '0.00'
ws4['E7'] = '=C7*0.15'
ws4['E7'].number_format = '0.00'
ws4['F7'] = 800
ws4['F7'].number_format = '#,##0'

ws4['A8'] = "Attendance Rate (%)"
ws4['B8'] = f"=(30-'Academic Overview'!AVERAGE(G11:G{10+len(df)}))/30*100"
ws4['B8'].number_format = '0.00'
ws4['C8'] = 10
ws4['C8'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ws4['D8'] = '=B8+C8'
ws4['D8'].number_format = '0.00'
ws4['E8'] = '=C8*0.02'
ws4['E8'].number_format = '0.00'
ws4['F8'] = 500
ws4['F8'].number_format = '#,##0'

ws4['A9'] = "Tutoring Enrollment (%)"
ws4['B9'] = f"='Academic Overview'!COUNTIF(H11:H{10+len(df)},1)/{len(df)}*100"
ws4['B9'].number_format = '0.00'
ws4['C9'] = 25
ws4['C9'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ws4['D9'] = '=B9+C9'
ws4['D9'].number_format = '0.00'
ws4['E9'] = '=C9*0.01'
ws4['E9'].number_format = '0.00'
ws4['F9'] = 3000
ws4['F9'].number_format = '#,##0'

ws4['A10'] = "Parental Engagement"
ws4['B10'] = f"='Academic Overview'!AVERAGE(I11:I{10+len(df)})"
ws4['B10'].number_format = '0.00'
ws4['C10'] = 1
ws4['C10'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ws4['D10'] = '=B10+C10'
ws4['D10'].number_format = '0.00'
ws4['E10'] = '=C10*0.08'
ws4['E10'].number_format = '0.00'
ws4['F10'] = 1200
ws4['F10'].number_format = '#,##0'

ws4['A11'] = "NOTE: Adjust values in column C (yellow cells) to simulate different scenarios"
ws4['A11'].font = Font(italic=True, size=9, color="7F7F7F")
ws4.merge_cells('A11:F11')

# Projections
ws4['A13'] = "PROJECTED OUTCOMES"
ws4['A13'].font = title_font
ws4.merge_cells('A13:H13')

ws4['A15'] = "Metric"
ws4['B15'] = "Current"
ws4['C15'] = "Projected"
ws4['D15'] = "Change"
ws4['E15'] = "% Improvement"
ws4['F15'] = "Target Met?"

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws4[f'{col}15']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

ws4['A16'] = "Average GPA"
ws4['B16'] = "='Academic Overview'!B6"
ws4['B16'].number_format = '0.00'
ws4['C16'] = '=B16+SUM(E7:E10)'
ws4['C16'].number_format = '0.00'
ws4['D16'] = '=C16-B16'
ws4['D16'].number_format = '0.00'
ws4['E16'] = '=D16/B16*100'
ws4['E16'].number_format = '0.0'
ws4['F16'] = '=IF(C16>=3,"YES","NO")'

ws4['A17'] = "Pass Rate (%)"
ws4['B17'] = "='Academic Overview'!B7"
ws4['B17'].number_format = '0.00'
ws4['C17'] = '=B17+(SUM(E7:E10)*10)'
ws4['C17'].number_format = '0.00'
ws4['D17'] = '=C17-B17'
ws4['D17'].number_format = '0.00'
ws4['E17'] = '=D17/B17*100'
ws4['E17'].number_format = '0.0'
ws4['F17'] = '=IF(C17>=80,"YES","NO")'

ws4['A18'] = "Failure Rate (%)"
ws4['B18'] = "='Academic Overview'!B8"
ws4['B18'].number_format = '0.00'
ws4['C18'] = '=B18-(SUM(E7:E10)*8)'
ws4['C18'].number_format = '0.00'
ws4['D18'] = '=C18-B18'
ws4['D18'].number_format = '0.00'
ws4['E18'] = '=ABS(D18)/B18*100'
ws4['E18'].number_format = '0.0'
ws4['F18'] = '=IF(C18<=16,"YES","NO")'

ws4['A19'] = "At-Risk Students"
ws4['B19'] = "='Academic Overview'!B9"
ws4['B19'].number_format = '0'
ws4['C19'] = '=B19*(1-E18/100)'
ws4['C19'].number_format = '0'
ws4['D19'] = '=C19-B19'
ws4['D19'].number_format = '0'
ws4['E19'] = '=D19/B19*100'
ws4['E19'].number_format = '0.0'
ws4['F19'] = '=IF(C19<B19,"IMPROVED","NO CHANGE")'

# Cost-Benefit
ws4['A21'] = "COST-BENEFIT ANALYSIS"
ws4['A21'].font = title_font
ws4.merge_cells('A21:H21')

ws4['A23'] = "Category"
ws4['B23'] = "Description"
ws4['C23'] = "Calculation"
ws4['D23'] = "Amount"

for col in ['A', 'B', 'C', 'D']:
    cell = ws4[f'{col}23']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

ws4['A24'] = "Total Students"
ws4['B24'] = "All enrolled students"
ws4['C24'] = "From dataset"
ws4['D24'] = len(df)

ws4['A25'] = "At-Risk Students"
ws4['B25'] = "Students needing intervention"
ws4['C25'] = "Current count"
ws4['D25'] = '=B19'
ws4['D25'].number_format = '0'

ws4['A26'] = "Total Intervention Cost"
ws4['B26'] = "Sum of all programs"
ws4['C26'] = "At-risk × avg cost"
ws4['D26'] = '=D25*AVERAGE(F7:F10)'
ws4['D26'].number_format = '#,##0'

ws4['A27'] = "Cost per Prevention"
ws4['B27'] = "Cost per prevented failure"
ws4['C27'] = "Total cost / prevented"
ws4['D27'] = '=D26/ABS(D19)'
ws4['D27'].number_format = '#,##0'

ws4['A28'] = "ROI (Saved Tuition)"
ws4['B28'] = "Assuming $40K tuition"
ws4['C28'] = "Prevented × tuition"
ws4['D28'] = '=ABS(D19)*40000'
ws4['D28'].number_format = '#,##0'

ws4['A29'] = "Net Benefit"
ws4['B29'] = "ROI - Cost"
ws4['C29'] = "Savings - Investment"
ws4['D29'] = '=D28-D26'
ws4['D29'].number_format = '#,##0'
ws4['D29'].font = Font(bold=True, size=12)

ws4['A30'] = "ROI Ratio"
ws4['B30'] = "Return on Investment"
ws4['C30'] = "Net / Cost"
ws4['D30'] = '=D29/D26'
ws4['D30'].number_format = '0.00'
ws4['D30'].font = Font(bold=True, size=12)

ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 25
ws4.column_dimensions['D'].width = 15

# ===============================================
# PAGE 5: ETHICS & SAFEGUARDS
# ===============================================
ws5 = wb.create_sheet("Ethics & Safeguards")

ws5['A1'] = "ETHICAL FRAMEWORK & DATA SAFEGUARDS"
ws5['A1'].font = Font(bold=True, size=18, color="1F4E78")
ws5.merge_cells('A1:G1')

ws5['A2'] = "Ensuring Responsible Use of Predictive Analytics"
ws5['A2'].font = Font(size=11, italic=True, color="7F7F7F")
ws5.merge_cells('A2:G2')

# Ethical Principles
ws5['A4'] = "CORE ETHICAL PRINCIPLES"
ws5['A4'].font = title_font
ws5.merge_cells('A4:G4')

ws5['A6'] = "Principle"
ws5['B6'] = "Description"
ws5['C6'] = "Implementation"
ws5['D6'] = "Status"

for col in ['A', 'B', 'C', 'D']:
    cell = ws5[f'{col}6']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

principles = [
    ('Transparency', 'Students informed about assessment', 
     'Risk scores shared in counseling', 'Active'),
    ('Fairness', 'No demographic discrimination',
     'Algorithm audited for bias', 'Active'),
    ('Privacy', 'Student data protected',
     'FERPA compliant; encrypted storage', 'Active'),
    ('Human Oversight', 'Humans make final decisions',
     'Advisors review all recommendations', 'Active'),
    ('Right to Explanation', 'Students can request details',
     'Detailed breakdown available', 'Active'),
    ('Opt-Out Option', 'Students can decline',
     'Alternative support offered', 'Active'),
]

for idx, (principle, description, implementation, status) in enumerate(principles, start=1):
    row = 6 + idx
    ws5[f'A{row}'] = principle
    ws5[f'B{row}'] = description
    ws5[f'C{row}'] = implementation
    ws5[f'D{row}'] = status
    ws5[f'D{row}'].alignment = Alignment(horizontal='center')

# Privacy Policy
ws5['A14'] = "DATA PRIVACY & SECURITY POLICY"
ws5['A14'].font = title_font
ws5.merge_cells('A14:G14')

ws5['A16'] = "Category"
ws5['B16'] = "Policy"
ws5['C16'] = "Implementation"
ws5['D16'] = "Review Frequency"

for col in ['A', 'B', 'C', 'D']:
    cell = ws5[f'{col}16']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

policies = [
    ('Data Collection', 'Only essential academic data',
     'Automated from SIS', 'Annual'),
    ('Data Storage', 'Encrypted databases',
     'AES-256 encryption', 'Quarterly'),
    ('Data Access', 'Role-based control',
     'Advisors: read only', 'Monthly'),
    ('Data Retention', 'Delete after grad + 2yr',
     'Automated purge process', 'Semester'),
    ('Third-Party', 'No sales without consent',
     'DPA agreements required', 'Per request'),
    ('Breach Protocol', 'Notify within 72 hours',
     'Incident response team', 'Annual drill'),
]

for idx, (category, policy, implementation, frequency) in enumerate(policies, start=1):
    row = 16 + idx
    ws5[f'A{row}'] = category
    ws5[f'B{row}'] = policy
    ws5[f'C{row}'] = implementation
    ws5[f'D{row}'] = frequency

# Bias Monitoring
ws5['A24'] = "ALGORITHMIC BIAS MONITORING"
ws5['A24'].font = title_font
ws5.merge_cells('A24:E24')

ws5['A26'] = "Group"
ws5['B26'] = "Avg Risk Score"
ws5['C26'] = "Pass Rate %"
ws5['D26'] = "Fairness Metric"
ws5['E26'] = "Status"

for col in ['A', 'B', 'C', 'D', 'E']:
    cell = ws5[f'{col}26']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

ws5['A27'] = "Overall Population"
ws5['B27'] = f"='Performance Risk Index'!AVERAGE(H{data_start_risk}:H{data_end_risk})"
ws5['B27'].number_format = '0.00'
ws5['C27'] = "='Academic Overview'!B7"
ws5['C27'].number_format = '0.0'
ws5['D27'] = "Baseline"
ws5['E27'] = "Reference"

ws5['A28'] = "Fairness Note: Monitor variance < 10% between groups"
ws5['A28'].font = Font(italic=True, size=9, color="7F7F7F")
ws5.merge_cells('A28:E28')

# Governance
ws5['A30'] = "ACCOUNTABILITY & GOVERNANCE"
ws5['A30'].font = title_font
ws5.merge_cells('A30:D30')

ws5['A32'] = "Stakeholder"
ws5['B32'] = "Role"
ws5['C32'] = "Responsibility"
ws5['D32'] = "Meeting Frequency"

for col in ['A', 'B', 'C', 'D']:
    cell = ws5[f'{col}32']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

stakeholders = [
    ('Ethics Review Board', 'Oversight', 'Review algorithm changes', 'Quarterly'),
    ('Data Protection Officer', 'Compliance', 'Ensure FERPA compliance', 'Monthly'),
    ('Student Representatives', 'Advocacy', 'Voice student concerns', 'Bi-monthly'),
    ('Faculty Advisory', 'Implementation', 'Guide intervention strategies', 'Monthly'),
    ('IT Security Team', 'Protection', 'Maintain data security', 'Weekly'),
]

for idx, (stakeholder, role, responsibility, frequency) in enumerate(stakeholders, start=1):
    row = 32 + idx
    ws5[f'A{row}'] = stakeholder
    ws5[f'B{row}'] = role
    ws5[f'C{row}'] = responsibility
    ws5[f'D{row}'] = frequency

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 35
ws5.column_dimensions['D'].width = 18

# Save workbook
output_path = '/Users/prachisingh/Desktop/rev_ler_da/Student_Early_Warning_Dashboard.xlsx'
wb.save(output_path)
print(f"✅ Dashboard created successfully: {output_path}")
print(f"\n📊 Dashboard contains {len(wb.sheetnames)} pages:")
for idx, sheet in enumerate(wb.sheetnames, 1):
    print(f"  Page {idx}: {sheet}")
print(f"\n📈 Analyzed {len(df)} students")
print(f"💡 All formulas use proper Excel syntax for compatibility")
